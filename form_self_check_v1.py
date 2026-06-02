# -*- coding: utf-8 -*-
"""
SPReAD 様式0/2/3/4 セルフチェックツール
（PDF 専用・Jupyter Notebook／コマンドライン 対応）

本ツールは、第2回 SPReAD 公募の応募書類 PDF を確認し、
結果を Microsoft Excel ファイルとして出力します。

■ ライセンス・利用条件
本ツールは、SPReAD 第2回公募の応募者向けに配布するセルフチェック用ツールです。
応募者本人、応募者所属機関による応募書類の確認を目的とした利用に限ります。

・本ツールの商用利用は禁止します。
・ご自身の利用の範囲内での改変は可能ですが、改変したものを第三者へ配布することは禁止します。
・本ツールは、必ず公式配布元から入手してください。第三者により再配布されたものの利用は認められません。

本ツールの使用により生じた損害について、文部科学省は責任を負いません。

■ 対象ファイル
- PDF のみを対象とします。
- Microsoft Word・画像・Microsoft Excel ファイルは対象外です。
- 手書きのスキャン PDFは文字を抽出できないため、「判断不可（PDF スキャン不可につき判断不可）」として出力します。

■ 出力される Microsoft Excel
以下の2つのシートを作成します。
- 判定サマリー
- 詳細チェック

■ 必要なライブラリ
初回に一度、以下を実行してください。
    pip install pdfminer.six pypdf openpyxl

■ 実行方法（コマンドライン）
    python form_self_check.py

対象フォルダを指定する場合：
    python form_self_check.py "C:\対象PDFフォルダ"

出力ファイル名も指定する場合：
    python form_self_check.py "C:\対象PDFフォルダ" --xlsx "C:\結果.xlsx"

■ 実行方法（Jupyter Notebook）
    from form_self_check import run
    run(r"C:\対象PDFフォルダ")

■ 設定の変更
通常は、本ファイル上部の INPUT_PATH／OUTPUT_FILE／RECURSIVE のみを変更します。
これら以外はツール内部の判定仕様です。
"""
from __future__ import annotations

import argparse
import json
import re
import sys
import unicodedata
from dataclasses import dataclass, asdict
from datetime import date, datetime
from numbers import Number
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

# ============================================================
# 設定（実行前に確認・変更する項目）
# ============================================================
# INPUT_PATH  : 確認対象。PDF を含むフォルダ、または個々の PDF ファイルのパスを指定します。
#               既定値 Path(".") は、本ファイルを実行したフォルダを対象とします。
#               例) Path(r"C:\Users\xxx\Desktop\応募書類")
#               ※ Windows のパスは、先頭に r を付けて r"..." の形式で指定してください。
# OUTPUT_FILE : 結果ファイルの名称。実行のたびに末尾へ日時が付与されるため、前回の結果は上書きされません。
#               別の場所へ保存する場合は、フルパスで指定してください。
# RECURSIVE   : True にすると、指定フォルダ配下のサブフォルダも含めて確認します。
#               提出者ごとのフォルダをまとめた親フォルダを指定し、機関側で一括確認する際にご利用ください
#               この場合、提出者フォルダの直下に PDF を配置してください。

INPUT_PATH = Path(".")
OUTPUT_FILE = Path("同意確認書_セルフチェック結果.xlsx")
RECURSIVE = False


# ============================================================
# これより下は変更しないでください。
VALID_CONFIRM_DATE_START = date(2026, 6, 2)
VALID_CONFIRM_DATE_END = date(2026, 7, 3)

SCAN_PDF_STATUS = "判断不可（PDFスキャン不可につき判断不可）"
MIN_EXTRACTED_TEXT_CHARS = 50

def get_timestamped_output_file() -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return OUTPUT_FILE.with_name(f"同意確認書_セルフチェック結果_{timestamp}.xlsx")

CONTROL_CHAR_RE = re.compile(r"[\x00-\x1f\x7f]")
# Excel セルに書き込めない制御文字（openpyxl が拒否する範囲）。タブ・改行・復帰は保持する。
ILLEGAL_CELL_CHAR_RE = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]")
MAX_EXCEL_CELL_LENGTH = 32767

def validate_path_string(path_value: str | Path) -> Path:
    """ユーザー指定パスの基本検証を行う。"""
    raw = str(path_value).strip()
    if not raw:
        raise ValueError("パスが空です")
    if CONTROL_CHAR_RE.search(raw):
        raise ValueError("不正な制御文字を含むパスが指定されました")

    p = Path(raw).expanduser()
    if any(part == ".." for part in p.parts):
        raise ValueError("親ディレクトリ参照（..）を含むパスは指定できません")
    return p

def safe_existing_path(path_value: str | Path) -> Path:
    """入力対象パスを検証し、存在する絶対パスとして返す。"""
    p = validate_path_string(path_value).resolve(strict=False)
    if not p.exists():
        raise FileNotFoundError(f"対象が見つかりません: {p}")
    return p

def safe_output_path(path_value: str | Path, expected_suffix: str | None = None) -> Path:
    """出力先パスを検証し、絶対パスとして返す。"""
    p = validate_path_string(path_value).resolve(strict=False)
    if expected_suffix and p.suffix.lower() != expected_suffix.lower():
        raise ValueError(f"出力ファイルの拡張子は {expected_suffix} にしてください: {p}")
    return p

def safe_text_for_output(value: object) -> object:
    """Excel/CSV 出力用に外部入力（PDF抽出文字列など）を安全化する。

    - Excel セルに書き込めない制御文字を除去する（これを行わないと
      openpyxl が IllegalCharacterError で停止し、一括実行時に結果が
      1件も出力されなくなる）
    - HTML/スクリプト断片として解釈され得る < > をエスケープする
    - Excel/CSV 数式インジェクションにつながる先頭文字を無害化する
    - Excel セル長の上限を超えないようにする

    ※ 様式1ツール（research_plan_self_check.py）の同名関数と挙動をそろえている。
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return value
    if isinstance(value, Number):
        return int(value) if isinstance(value, float) and value.is_integer() else value
    if isinstance(value, (datetime, date)):
        return value

    text = str(value)
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = ILLEGAL_CELL_CHAR_RE.sub("", text)
    text = text.replace("<", "&lt;").replace(">", "&gt;")

    formula_probe = text.lstrip(" \t\n")
    if formula_probe.startswith(("=", "+", "@")) or re.match(r"^-[A-Za-z0-9_().]", formula_probe):
        text = "'" + text

    return text[:MAX_EXCEL_CELL_LENGTH]

CHECKED_MARKS = {"☒", "■", "✅", "✓", "✔", "[x]", "[X]", "true", "TRUE"}
UNCHECKED_MARKS = {"☐", "□", "[ ]", "false", "FALSE"}
PLACEHOLDER_RE = re.compile(r"^(?:[_＿\s]+|X{2,}|x{2,}|○+|〇+|あ+|テスト|未定|N/?A|なし)$", re.I)

FILENAME_RULES = {
    "様式0": "第2回_様式0_申請様式チェックリスト",
    "様式2": "第2回_様式2_審査手法及び応募に係る情報の取扱い等に関する同意確認書",
    "様式3": "第2回_様式3_学生応募の同意確認書",
    "様式4": "第2回_様式4_指導教員等の同意確認書",
}

DETAIL_MASTER: Dict[str, List[Tuple[str, str]]] = {
    "様式0": [
        ("対象書類の確認", "第2回目の公募書類を使用しているか"),
        ("対象書類の確認", "様式のファイル名が指定どおりになっている"),
        ("1．応募書類の確認", "チェック項目: 必要書類がすべて提出されている"),
        ("1．応募書類の確認", "チェック項目: 研究インテグリティの確保に係る誓約状況をはじめとして、e-Rad に必要な研究者情報が登録されている"),
        ("1．応募書類の確認", "チェック項目: 様式0のファイル名指定の確認"),
        ("1．応募書類の確認", "チェック項目: 様式2のファイル名指定の確認"),
        ("1．応募書類の確認", "チェック項目: 様式3のファイル名指定の確認"),
        ("1．応募書類の確認", "チェック項目: 様式4のファイル名指定の確認"),
        ("2．提出形式の確認", "チェック項目: 研究計画調書がExcel形式で作成されている"),
        ("2．提出形式の確認", "チェック項目: 研究計画調書以外の様式については、PDF形式に変換されている"),
        ("2．提出形式の確認", "チェック項目:ファイル名が指定どおりとなっている"),
        ("2．提出形式の確認", "チェック項目: e-Rad所属機関コードが半角数字で記載されている"),
        ("2．提出形式の確認", "チェック項目: ローマ字氏名は半角英字で「姓・名」の順になっている"),
        ("3．記載内容の確認", "チェック項目:各様式に記入漏れがない"),
        ("3．記載内容の確認", "チェック項目: 研究計画調書については、未入力セル（薄いオレンジ色）が必要箇所で解消されている"),
        ("3．記載内容の確認", "チェック項目: 研究計画調書内の文字数制限のある項目について、制限内で記載されている"),
        ("3．記載内容の確認", "チェック項目:  研究計画調書に記載された研究経費は公募要領に定められた範囲内であり、計算ミスがない"),
        ("3．記載内容の確認", "チェック項目: 研究計画調書には、所属機関から発行されている研究代表者本人のメールアドレスが記載されている"),
        ("最終確認", "チェック項目: 上記1～3をすべて確認しました"),
        ("確認者情報", "氏名"),
        ("確認者情報", "所属機関"),
        ("確認者情報", "部署"),
        ("確認者情報", "確認日"),
    ],
    "様式2": [
        ("対象書類の確認", "第2回目の公募書類を使用しているか"),
        ("対象書類の確認", "様式のファイル名が指定どおりになっている"),
        ("最終確認", "上記項目を全て確認しました"),
        ("確認者情報", "氏名"),
        ("確認者情報", "所属機関"),
        ("確認者情報", "部署"),
        ("確認者情報", "確認日"),
    ],
    "様式3": [
        ("対象書類の確認", "第2回目の公募書類を使用しているか"),
        ("対象書類の確認", "様式のファイル名が指定どおりになっている"),
        ("最終確認", "上記項目を全て確認しました"),
        ("確認者情報", "氏名"),
        ("確認者情報", "所属機関"),
        ("確認者情報", "部署"),
        ("確認者情報", "身分・課程等／学年"),
        ("確認者情報", "確認日"),
    ],
    "様式4": [
        ("対象書類の確認", "第2回目の公募書類を使用しているか"),
        ("対象書類の確認", "様式のファイル名が指定どおりになっている"),
        ("最終確認", "上記項目を全て確認しました"),
        ("確認者情報", "氏名"),
        ("確認者情報", "所属機関"),
        ("確認者情報", "部署"),
        ("確認者情報", "役職"),
        ("確認者情報", "記載欄（所見）"),
        ("確認者情報", "確認日"),
    ],
}

OK_STATUSES = {"OK", "記入済み", "チェック済み"}
NON_BLOCKING_STATUSES = {"未チェック（学生の場合必須）"}
NG_PREFIXES = ("要確認", "未記入", "未チェック", "読込エラー", "様式不明", "日付要確認", "判定不可", SCAN_PDF_STATUS)

@dataclass
class DetailRow:
    ファイル名: str
    タブ名: str
    大項目: str
    項目: str
    判定: str
    参照セル: str = ""
    入力値: str = ""

def normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKC", str(value))
    text = text.replace("\u3000", " ").replace("\xa0", " ")
    text = re.sub(r"[\t\r]+", " ", text)
    text = re.sub(r"[ ]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()

def compact_for_match(value: object) -> str:
    """判定用に、全角/半角差・空白・改行・タブを無視した文字列へ変換する。"""
    if value is None:
        return ""
    text = unicodedata.normalize("NFKC", str(value))
    return re.sub(r"\s+", "", text)

def flexible_phrase_regex(phrase: str) -> str:
    """語句の各文字間に任意の空白が入っても一致する正規表現を作る。"""
    compact = compact_for_match(phrase)
    return r"\s*".join(re.escape(ch) for ch in compact)

def rfind_flexible(text: str, phrases: Iterable[str]) -> int:
    """空白差分を無視して、複数候補の最後の出現位置を返す。"""
    last = -1
    for phrase in phrases:
        pattern = flexible_phrase_regex(phrase)
        for m in re.finditer(pattern, text, flags=re.S):
            if m.start() > last:
                last = m.start()
    return last

def extract_pdf_text(path: Path) -> str:
    errors = []
    # 1) pdfminer.six（MIT）でテキスト抽出
    try:
        from pdfminer.high_level import extract_text as pdfminer_extract_text
        text = normalize_text(pdfminer_extract_text(str(path)) or "")
        if text.strip():
            return text
    except Exception as e1:
        errors.append(f"pdfminer.six: {e1}")
    # 2) pypdf（BSD）でフォールバック抽出
    try:
        from pypdf import PdfReader
        reader = PdfReader(str(path))
        return normalize_text("\n".join(page.extract_text() or "" for page in reader.pages))
    except Exception as e2:
        errors.append(f"pypdf: {e2}")
    raise RuntimeError(
        "PDFテキスト抽出に失敗しました。pdfminer.six または pypdf をインストールしてください。"
        f" 詳細: {'; '.join(errors)}"
    )

def extract_text(path: str | Path) -> str:
    p = Path(path)
    if p.suffix.lower() != ".pdf":
        raise ValueError(f"PDF専用です。未対応の拡張子: {p.suffix}")
    return extract_pdf_text(p)

def normalize_filename_for_rule(value: object) -> str:
    text = unicodedata.normalize("NFKC", str(value))
    text = text.replace(" ", "").replace("　", "")
    return text.strip()

def detect_form_type_from_filename(file_path: Path) -> str:
    name = normalize_filename_for_rule(file_path.stem)
    for form, prefix in FILENAME_RULES.items():
        if name.startswith(normalize_filename_for_rule(prefix) + "_"):
            return form
    m = re.search(r"様式([0234])", name)
    return f"様式{m.group(1)}" if m else "不明"

def detect_form_type(text: str, filename: str = "") -> str:
    normalized = unicodedata.normalize("NFKC", text)
    m = re.search(r"第\s*\d+\s*回\s*[_＿]?\s*様式\s*([0234])", normalized)
    if m:
        return f"様式{m.group(1)}"
    target = compact_for_match(text)
    if compact_for_match("申請様式チェックリスト") in target:
        return "様式0"
    if compact_for_match("審査手法及び応募に係る情報") in target:
        return "様式2"
    if compact_for_match("学生応募の同意確認書") in target:
        return "様式3"
    if compact_for_match("指導教員等の同意確認書") in target:
        return "様式4"
    return detect_form_type_from_filename(Path(filename)) if filename else "不明"

def judge_round_document(text: str, expected_form: str = "") -> Tuple[str, str]:
    # 「第2回_様式X」は文書のヘッダーにある前提で、先頭付近のみを見る。
    # 本文（特に様式0のチェックリストには様式2/3/4への参照が並ぶ）を拾って
    # 様式番号を取り違えないよう、全文フォールバックは行わない。
    normalized = unicodedata.normalize("NFKC", text)
    head = normalized[:500]
    pattern = r"第\s*2\s*回\s*[_＿]?\s*様式\s*([0234])"
    m = re.search(pattern, head)
    if not m:
        return "要確認", ""
    form_no = m.group(1)
    # 判定済みの様式（expected_form）と見出しの様式番号が食い違う場合は要確認。
    if expected_form and f"様式{form_no}" != expected_form:
        return "要確認", f"第2回_様式{form_no}"
    return "OK", f"第2回_様式{form_no}"

def split_filename_parts(file_path: Path, form: str) -> Tuple[Optional[str], Optional[str], str]:
    """prefix_erad_romaji の erad と romaji を取り出す。構造が大きく崩れている場合は None。"""
    if form not in FILENAME_RULES:
        return None, None, "ファイル名形式"
    stem = normalize_filename_for_rule(file_path.stem)
    prefix = normalize_filename_for_rule(FILENAME_RULES[form])
    if not stem.startswith(prefix + "_"):
        return None, None, "ファイル名形式"
    rest = stem[len(prefix) + 1:]
    parts = rest.split("_", 1)
    if len(parts) != 2:
        return None, None, "ファイル名形式"
    return parts[0], parts[1], ""

def judge_file_name(file_path: Path, form: str = "") -> Tuple[str, str, str]:
    """ファイル名を判定する。"""
    if file_path.name.startswith("~$"):
        return "要確認（ファイル名形式）", file_path.name, "一時ファイルです"

    target_form = form if form in FILENAME_RULES else detect_form_type_from_filename(file_path)
    if target_form not in FILENAME_RULES:
        return "要確認（ファイル名形式）", file_path.name, "様式0/2/3/4のファイル名ルールに該当しません"

    erad_code, romaji_name, structure_error = split_filename_parts(file_path, target_form)
    if structure_error:
        return "要確認（ファイル名形式）", file_path.name, f"期待形式: {FILENAME_RULES[target_form]}_e-Rad所属機関コード_ローマ字氏名"

    errors: List[str] = []
    if erad_code is None or romaji_name is None:
        return "要確認（ファイル名形式）", file_path.name, "ファイル名解析エラー"

    if not erad_code.isdigit() or len(erad_code) not in {3, 4, 7, 10}:
        errors.append("e-Rad所属機関コード")

    if not re.fullmatch(r"[A-Za-z][A-Za-z0-9._-]*", romaji_name):
        errors.append("ローマ字氏名")

    if errors:
        return f"要確認（{'、'.join(errors)}）", file_path.name, "、".join(errors)
    return "OK", file_path.name, "ファイル名ルールに一致"

def is_meaningful(value: Optional[str]) -> Tuple[bool, str]:
    if value is None:
        return False, "項目が見つかりません"
    v = value.strip()
    if not v:
        return False, "空欄です"
    if PLACEHOLDER_RE.fullmatch(v):
        return False, "プレースホルダーまたは仮入力の可能性があります"
    return True, ""

def label_regex(label: str) -> str:
    parts = []
    for ch in label:
        if ch in {"/", "／"}:
            parts.append(r"[/／]")
        else:
            parts.append(re.escape(ch))
    return "".join(parts)

def signature_block(text: str, form: str) -> str:
    m_end = re.search(flexible_phrase_regex("※補足事項"), text, flags=re.S)
    body = text[:m_end.start()] if m_end else text
    anchors = {
        "様式0": ["研究代表者 氏名", "研究代表者氏名", "上記1~3をすべて確認しました", "上記1～3をすべて確認しました"],
        "様式2": ["以上", "上記項目を全て確認しました"],
        "様式3": ["研究代表者 氏名", "研究代表者氏名", "上記項目を全て確認しました"],
        "様式4": ["指導教員等 氏名", "指導教員等氏名", "上記項目を全て確認しました"],
    }.get(form, [])
    positions = [rfind_flexible(body, [a]) for a in anchors]
    positions = [pos for pos in positions if pos >= 0]
    return body[min(positions):] if positions else body

def get_value_after_label(text: str, label: str) -> Optional[str]:
    labels = ["氏名", "所属機関", "部署", "役職", "身分・課程等／学年", "身分・課程等/学年", "確認日"]
    stop_patterns = [label_regex(x) for x in labels if normalize_text(x) != normalize_text(label)]
    stop = "|".join(stop_patterns)
    pattern = rf"{label_regex(label)}\s*[:：]?\s*(.*?)(?=\n|(?:{stop})\s*[:：]?|$)"
    matches = list(re.finditer(pattern, text, flags=re.S))
    if not matches:
        return None
    candidates: List[str] = []
    for m in matches:
        value = m.group(1)
        value = re.sub(r"[＿_]+", "", value)
        value = re.sub(r"\s+", " ", value).strip(" 　:：")
        if value:
            candidates.append(value)
    return candidates[-1] if candidates else ""

def parse_japanese_date(value: str) -> Optional[date]:
    text = normalize_text(value)
    text = text.replace("＿", "").replace("_", "")
    m = re.search(r"(20\d{2})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日", text)
    if m:
        y, mo, d = map(int, m.groups())
        try:
            return date(y, mo, d)
        except ValueError:
            return None
    m = re.search(r"令和\s*(\d+)\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日", text)
    if m:
        ry, mo, d = map(int, m.groups())
        try:
            return date(2018 + ry, mo, d)
        except ValueError:
            return None
    return None

def extract_confirm_date_value(block: str) -> str:
    v = get_value_after_label(block, "確認日")
    if v:
        candidate = "確認日 " + v
        d = parse_japanese_date(candidate)
        if d:
            return f"{d.year}年{d.month}月{d.day}日"
        return v

    cleaned = block.replace("＿", "").replace("_", "")
    m = re.search(r"(20\d{2})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日", cleaned)
    if m:
        y, mo, d = map(int, m.groups())
        return f"{y}年{mo}月{d}日"
    m = re.search(r"令和\s*(\d+)\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日", cleaned)
    if m:
        ry, mo, d = map(int, m.groups())
        return f"令和{ry}年{mo}月{d}日"
    return ""

def judge_confirm_date(value: str) -> str:
    dt = parse_japanese_date(value)
    if not dt:
        return "未記入"
    if VALID_CONFIRM_DATE_START <= dt <= VALID_CONFIRM_DATE_END:
        return "記入済み"
    return "日付要確認"

def is_scanned_pdf_text(text: str) -> bool:
    """OCRされていないスキャンPDFの可能性を判定する。"""
    cleaned = re.sub(r"\s+", "", normalize_text(text))
    return len(cleaned) < MIN_EXTRACTED_TEXT_CHARS

def extract_opinion_value(text: str) -> str:
    """様式4の記載欄（所見）を抽出する。"""
    m_end = re.search(flexible_phrase_regex("※補足事項"), text, flags=re.S)
    body = text[:m_end.start()] if m_end else text

    start_match = None
    for phrase in ["必ずご記入ください", "記載欄", "所見"]:
        m = re.search(flexible_phrase_regex(phrase), body, flags=re.S)
        if m and (start_match is None or m.end() > start_match.end()):
            start_match = m

    if not start_match:
        return ""

    rest = body[start_match.end():]

    end_positions = []
    for phrase in ["上記項目を全て確認しました", "上記項目をすべて確認しました", "指導教員等", "氏名", "所属機関", "部署", "役職", "確認日"]:
        m = re.search(flexible_phrase_regex(phrase), rest, flags=re.S)
        if m:
            end_positions.append(m.start())

    m_date = re.search(r"(?:20\d{2}|令和\s*\d+)\s*[＿_ ]*年\s*[＿_ ]*\d{1,2}\s*月\s*[＿_ ]*\d{1,2}\s*[＿_ ]*日", rest)
    if m_date:
        end_positions.append(m_date.start())

    opinion = rest[:min(end_positions)] if end_positions else rest
    opinion = re.sub(r"[＿_]+", "", opinion)
    opinion = re.sub(r"\s+", "", opinion)
    return opinion.strip(" 　:：。")

def judge_opinion(value: str) -> str:
    if not value:
        return "未記入"
    if PLACEHOLDER_RE.fullmatch(value):
        return "未記入"
    return "記入済み"

def normalize_for_search(value: object) -> str:
    """検索・判定用に、全角/半角差と空白差分を吸収する。"""
    return compact_for_match(value)

def build_item_candidates(item: str) -> List[str]:
    """詳細チェックの項目名から、PDF本文で探す候補語を作る。"""
    phrase = item.replace("チェック項目:", "").strip()
    candidates = [phrase]

    if "様式0のファイル名指定" in item:
        candidates.extend([
            "第2回_様式0_申請様式チェックリスト",
            "様式0_申請様式チェックリスト",
        ])
    elif "様式1のファイル名指定" in item:
        candidates.extend([
            "第2回_様式1_研究計画調書",
            "様式1_研究計画調書",
        ])
    elif "様式2のファイル名指定" in item:
        candidates.extend([
            "第2回_様式2_審査手法及び応募に係る情報の取扱い等に関する同意確認書",
            "様式2_審査手法及び応募に係る情報の取扱い等に関する同意確認書",
        ])
    elif "様式3のファイル名指定" in item:
        candidates.extend([
            "第2回_様式3_学生応募の同意確認書",
            "様式3_学生応募の同意確認書",
        ])
    elif "様式4のファイル名指定" in item:
        candidates.extend([
            "第2回_様式4_指導教員等の同意確認書",
            "様式4_指導教員等の同意確認書",
        ])

    if "上記1～3" in phrase or "上記1~3" in phrase:
        candidates.extend([
            "上記1～3をすべて確認しました",
            "上記1~3をすべて確認しました",
            "上記 1～3 をすべて確認しました",
        ])

    return list(dict.fromkeys([c for c in candidates if c]))

def find_phrase_match(text: str, item: str) -> Optional[re.Match]:
    """空白差分を許容して、元テキスト上のマッチ位置を返す。"""
    for candidate in build_item_candidates(item):
        pattern = flexible_phrase_regex(candidate)
        m = re.search(pattern, text, flags=re.S)
        if m:
            return m
    return None

def nearest_checkbox_status(window: str, phrase_local_start: int) -> Optional[str]:
    """対象行付近のチェック記号から状態を判定する。"""
    marks: List[Tuple[int, str]] = []
    for mark in CHECKED_MARKS:
        for m in re.finditer(re.escape(mark), window):
            marks.append((m.start(), "チェック済み"))
    for mark in UNCHECKED_MARKS:
        for m in re.finditer(re.escape(mark), window):
            marks.append((m.start(), "未チェック"))

    if not marks:
        return None

    before = [(pos, status) for pos, status in marks if pos <= phrase_local_start]
    if before:
        pos, status = min(before, key=lambda x: phrase_local_start - x[0])
        return status

    pos, status = min(marks, key=lambda x: abs(x[0] - phrase_local_start))
    return status

def has_checked_near(text: str, item: str) -> Tuple[str, str]:
    """チェック項目のチェック状態を判定する。"""
    m = find_phrase_match(text, item)
    if not m:
        return "判定不可", ""

    line_start = text.rfind("\n", 0, m.start()) + 1
    line_end = text.find("\n", m.end())
    if line_end < 0:
        line_end = len(text)

    window_start = max(0, line_start - 5)
    window_end = min(len(text), line_end + 5)
    window = text[window_start:window_end]
    phrase_local_start = m.start() - window_start

    status = nearest_checkbox_status(window, phrase_local_start)
    if status:
        return status, ""

    return "判定不可", ""

def has_checked_confirmation(text: str) -> bool:
    normalized = normalize_for_search(text)
    candidates = [
        "上記項目を全て確認しました",
        "上記項目をすべて確認しました",
        "上記1～3をすべて確認しました",
        "上記1~3をすべて確認しました",
    ]
    for candidate in candidates:
        idx = normalized.find(normalize_for_search(candidate))
        if idx >= 0:
            window = normalized[max(0, idx - 30): idx + len(normalize_for_search(candidate)) + 30]
            if any(mark in window for mark in CHECKED_MARKS):
                return True
    for m in re.finditer(r"上記.{0,40}確認しました", normalized):
        window = normalized[max(0, m.start() - 30): m.end() + 30]
        if any(mark in window for mark in CHECKED_MARKS):
            return True
    return False

def is_blocking_status(status: str) -> bool:
    if status in OK_STATUSES:
        return False
    if status in NON_BLOCKING_STATUSES:
        return False
    if status.startswith("要確認"):
        return True
    return status.startswith(NG_PREFIXES)

def normalize_summary_status(status: str) -> str:
    return "要確認" if is_blocking_status(status) else "OK"

def detail_row(file_path: Path, form: str, major: str, item: str, status: str, ref: str = "", value: object = "") -> DetailRow:
    return DetailRow(
        ファイル名=file_path.name,
        タブ名=form,
        大項目=major,
        項目=item,
        判定=status,
        参照セル=ref,
        入力値="" if value is None else str(value),
    )

def collect_files(input_path: str | Path) -> List[Path]:
    p = safe_existing_path(input_path)
    if p.is_file():
        return [p] if p.suffix.lower() == ".pdf" and not p.name.startswith("~$") else []
    files: List[Path] = []
    files.extend(p.rglob("*.pdf") if RECURSIVE else p.glob("*.pdf"))
    return sorted([x for x in files if x.is_file() and not x.name.startswith("~$")], key=lambda x: str(x).lower())

def detect_available_forms(files: List[Path]) -> set[str]:
    forms: set[str] = set()
    for f in files:
        try:
            forms.add(detect_form_type(extract_text(f), f.name))
        except Exception:
            forms.add(detect_form_type_from_filename(f))
    return forms

def build_detail_rows_for_file(file_path: Path, text: str, form: str, all_files: List[Path], available_forms: set[str]) -> List[DetailRow]:
    rows: List[DetailRow] = []
    block = signature_block(text, form)
    filename_status, filename_value, _filename_note = judge_file_name(file_path, form)
    round_status, round_value = judge_round_document(text, form)

    field_values: Dict[str, str] = {}
    for label in ["氏名", "所属機関", "部署", "役職", "身分・課程等／学年"]:
        field_values[label] = get_value_after_label(block, label) or ""
    field_values["確認日"] = extract_confirm_date_value(block)
    if form == "様式4":
        field_values["記載欄（所見）"] = extract_opinion_value(text)

    for major, item in DETAIL_MASTER.get(form, []):
        status = "判定不可"
        ref = ""
        value = ""

        if item == "第2回目の公募書類を使用しているか":
            status = round_status
            ref = "文書ヘッダ"
            value = round_value

        elif item == "様式のファイル名が指定どおりになっている":
            status, value, _ = judge_file_name(file_path, form)
            ref = "ファイル名"

        elif "様式3のファイル名指定の確認" in item and "様式3" not in available_forms:
            checked, _ = has_checked_near(text, item)
            status = checked if checked == "チェック済み" else "未チェック（学生の場合必須）"
            ref = "チェックボックス"
            value = ""

        elif "様式4のファイル名指定の確認" in item and "様式4" not in available_forms:
            checked, _ = has_checked_near(text, item)
            status = checked if checked == "チェック済み" else "未チェック（学生の場合必須）"
            ref = "チェックボックス"
            value = ""

        elif item.startswith("チェック項目:"):
            status, _ = has_checked_near(text, item)
            ref = "チェックボックス"
            value = ""

        elif item == "上記項目を全て確認しました":
            status = "チェック済み" if has_checked_confirmation(text) else "未チェック"
            ref = "確認チェック"
            value = ""

        elif item in {"氏名", "所属機関", "部署", "役職", "身分・課程等／学年"}:
            value = field_values.get(item, "")
            ok, _reason = is_meaningful(value)
            status = "記入済み" if ok else "未記入"
            ref = item

        elif item == "記載欄（所見）":
            value = field_values.get("記載欄（所見）", "")
            status = judge_opinion(value)
            ref = "記載欄（所見）"

        elif item == "確認日":
            value = field_values.get("確認日", "")
            status = judge_confirm_date(value)
            ref = "確認日"

        rows.append(detail_row(file_path, form, major, item, status, ref, value))

    unique: List[DetailRow] = []
    seen = set()
    for r in rows:
        key = (r.ファイル名, r.タブ名, r.大項目, r.項目)
        if key not in seen:
            seen.add(key)
            unique.append(r)
    return unique

def process_one_file(file_path: Path, all_files: List[Path], available_forms: set[str]) -> Tuple[Dict[str, str], List[DetailRow]]:
    summary: Dict[str, str] = {
        "ファイル名": file_path.name,
        "ファイルパス": str(file_path),
        "総合判定": "要確認",
        "指定されたファイル名になっているか": "要確認",
        "第2回公募書類を使用しているか": "要確認",
        "チェック項目": "要確認",
        "記入項目": "要確認",
        "要確認数": "0",
        "要確認項目": "",
        "エラー内容": "",
    }

    if file_path.suffix.lower() != ".pdf":
        summary["エラー内容"] = "PDFではありません"
        summary["要確認項目"] = "ファイル形式"
        summary["要確認数"] = "1"
        return summary, []

    try:
        text = extract_text(file_path)
        if is_scanned_pdf_text(text):
            summary["総合判定"] = SCAN_PDF_STATUS
            summary["指定されたファイル名になっているか"] = SCAN_PDF_STATUS
            summary["第2回公募書類を使用しているか"] = SCAN_PDF_STATUS
            summary["チェック項目"] = SCAN_PDF_STATUS
            summary["記入項目"] = SCAN_PDF_STATUS
            summary["要確認数"] = ""
            summary["要確認項目"] = ""
            summary["エラー内容"] = SCAN_PDF_STATUS
            return summary, [detail_row(file_path, "全体", "ファイル読込", "OCR判定", SCAN_PDF_STATUS, "PDFテキスト抽出", "")]

        form = detect_form_type(text, file_path.name)
        if form == "不明":
            summary["エラー内容"] = "様式0/2/3/4を判定できません"
            summary["要確認項目"] = "様式判定"
            summary["要確認数"] = "1"
            return summary, [detail_row(file_path, "不明", "対象書類の確認", "第2回目の公募書類を使用しているか", "要確認", "文書ヘッダ", "")]

        rows = build_detail_rows_for_file(file_path, text, form, all_files, available_forms)

        file_status, _file_value, _file_note = judge_file_name(file_path, form)
        round_status, _round_value = judge_round_document(text, form)
        summary["指定されたファイル名になっているか"] = normalize_summary_status(file_status)
        summary["第2回公募書類を使用しているか"] = normalize_summary_status(round_status)

        checkbox_rows = [r for r in rows if r.大項目 in {"1．応募書類の確認", "2．提出形式の確認", "3．記載内容の確認", "最終確認"}]
        required_rows = [r for r in rows if r.大項目 == "確認者情報"]

        checkbox_blocking = [r for r in checkbox_rows if is_blocking_status(r.判定)]
        required_blocking = [r for r in required_rows if is_blocking_status(r.判定)]

        summary["チェック項目"] = "要確認" if checkbox_blocking else "OK"
        summary["記入項目"] = "要確認" if required_blocking else "OK"

        blocking_rows = [r for r in rows if is_blocking_status(r.判定)]
        item_names = [r.項目 for r in blocking_rows]
        item_names = list(dict.fromkeys(item_names))
        summary["要確認数"] = str(len(item_names))
        summary["要確認項目"] = "、".join(item_names)
        summary["総合判定"] = "要確認" if item_names else "OK"
        return summary, rows

    except Exception as e:
        summary["エラー内容"] = str(e)
        summary["要確認項目"] = "ファイル読込"
        summary["要確認数"] = "1"
        return summary, [detail_row(file_path, "全体", "ファイル読込", "ファイル読込", "読込エラー", "", "")]

def check_folder(input_path: str | Path) -> Tuple[List[Dict[str, str]], List[Dict[str, str]]]:
    files = collect_files(input_path)
    # 1フォルダ＝1提出者として、フォルダ（直近の親）ごとに独立して処理する（学生判定が他フォルダの影響を受けないようにするため）。
    from collections import OrderedDict
    groups: "OrderedDict[Path, List[Path]]" = OrderedDict()
    for file_path in files:
        groups.setdefault(file_path.parent, []).append(file_path)

    summaries: List[Dict[str, str]] = []
    details_all: List[Dict[str, str]] = []
    for folder, folder_files in groups.items():
        available_forms = detect_available_forms(folder_files)
        for file_path in folder_files:
            summary, details = process_one_file(file_path, folder_files, available_forms)
            summary["フォルダ"] = folder.name
            summaries.append(summary)
            for d in details:
                row = asdict(d)
                row["フォルダ"] = folder.name
                details_all.append(row)
    return summaries, details_all

def check_file(input_path: str | Path) -> Tuple[Dict[str, str], List[Dict[str, str]]]:
    p = Path(input_path)
    files = [p]
    available_forms = detect_available_forms(files)
    summary, details = process_one_file(p, files, available_forms)
    summary["フォルダ"] = p.parent.name
    rows = []
    for d in details:
        row = asdict(d)
        row["フォルダ"] = p.parent.name
        rows.append(row)
    return summary, rows

def write_sheet(ws, headers: List[str], rows: List[Dict[str, str]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append([safe_text_for_output(row.get(h, "")) for h in headers])

def style_workbook(wb) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="1F4E78")
    ok_fill = PatternFill("solid", fgColor="E2F0D9")
    ng_fill = PatternFill("solid", fgColor="FCE4D6")
    warn_fill = PatternFill("solid", fgColor="FFF2CC")
    gray_fill = PatternFill("solid", fgColor="D9D9D9")
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border
                value = cell.value
                if value == "未チェック（学生の場合必須）":
                    cell.fill = gray_fill
                    cell.font = Font(color="808080")
                elif isinstance(value, str) and value.startswith("要確認"):
                    cell.fill = ng_fill
                elif value in {"未記入", "未チェック", "読込エラー", "様式不明", "日付要確認"}:
                    cell.fill = ng_fill
                elif value == SCAN_PDF_STATUS:
                    cell.fill = gray_fill
                    cell.font = Font(color="808080")
                elif value in {"判定不可"}:
                    cell.fill = warn_fill
                elif value in OK_STATUSES or value == "OK":
                    cell.fill = ok_fill
        for col_idx in range(1, ws.max_column + 1):
            header = ws.cell(1, col_idx).value or ""
            letter = get_column_letter(col_idx)
            if header in {"ファイルパス", "要確認項目", "入力値"}:
                width = 55
            elif header == "フォルダ":
                width = 24
            elif header == "ファイル名":
                width = 42
            elif header in {"大項目"}:
                width = 22
            elif header in {"項目"}:
                width = 62
            elif header in {"判定", "総合判定", "チェック項目", "記入項目"}:
                width = 24
            elif header in {"指定されたファイル名になっているか", "第2回公募書類を使用しているか"}:
                width = 28
            elif header in {"参照セル", "要確認数", "タブ名"}:
                width = 16
            else:
                width = min(max(len(str(header)) + 4, 14), 32)
            ws.column_dimensions[letter].width = width
        ws.row_dimensions[1].height = 36

def make_detail_sheet_title(folder_name: str, used: set) -> str:
    # Excelのシート名制約（31文字以内・ : \ / ? * [ ] 不可・重複不可）に合わせて整える。
    base = folder_name or "(フォルダなし)"
    base = re.sub(r"[:\\/?*\[\]]", "_", base)
    title = ("詳細_" + base)[:31]
    candidate = title
    n = 2
    while candidate.lower() in {u.lower() for u in used}:
        suffix = f"_{n}"
        candidate = title[: 31 - len(suffix)] + suffix
        n += 1
    used.add(candidate)
    return candidate

def build_output(summaries: List[Dict[str, str]], details: List[Dict[str, str]], output_file: str | Path) -> Path:
    from openpyxl import Workbook
    from collections import OrderedDict

    output_path = safe_output_path(output_file, ".xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "判定サマリー"

    summary_headers = [
        "フォルダ",
        "ファイル名",
        "総合判定",
        "指定されたファイル名になっているか",
        "第2回公募書類を使用しているか",
        "チェック項目",
        "記入項目",
        "要確認数",
        "要確認項目",
        "エラー内容",
    ]
    # 要確認数は数値としてExcelに出す（文字列だと並べ替え・集計ができないため）。
    summaries_for_sheet = []
    for s in summaries:
        s = dict(s)
        v = s.get("要確認数", "")
        if isinstance(v, str) and v.isdigit():
            s["要確認数"] = int(v)
        summaries_for_sheet.append(s)
    write_sheet(ws, summary_headers, summaries_for_sheet)

    # 詳細チェックはフォルダごとに別シートで出力する。
    detail_headers = ["ファイル名", "タブ名", "大項目", "項目", "判定", "参照セル", "入力値"]
    by_folder: "OrderedDict[str, List[Dict[str, str]]]" = OrderedDict()
    for d in details:
        by_folder.setdefault(d.get("フォルダ", ""), []).append(d)

    used_titles: set = set()
    for folder_name, rows in by_folder.items():
        ws_detail = wb.create_sheet(make_detail_sheet_title(folder_name, used_titles))
        detail_rows = [{h: r.get(h, "") for h in detail_headers} for r in rows]
        write_sheet(ws_detail, detail_headers, detail_rows)

    style_workbook(wb)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path

def write_json(summaries: List[Dict[str, str]], details: List[Dict[str, str]], output_file: str | Path) -> None:
    data = {"判定サマリー": summaries, "詳細チェック": details}
    output_path = safe_output_path(output_file, ".json")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

def write_csv(details: List[Dict[str, str]], output_file: str | Path) -> None:
    import csv
    headers = ["フォルダ", "ファイル名", "タブ名", "大項目", "項目", "判定", "参照セル", "入力値"]
    output_path = safe_output_path(output_file, ".csv")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        for d in details:
            writer.writerow({h: safe_text_for_output(d.get(h, "")) for h in headers})

def print_report(summaries: List[Dict[str, str]]) -> None:
    print(f"対象ファイル: {len(summaries)}件")
    for s in summaries:
        folder = s.get("フォルダ", "")
        prefix = f"{folder}/" if folder else ""
        print(f"[{s.get('総合判定')}] {prefix}{s.get('ファイル名')} | 要確認数: {s.get('要確認数')} | 要確認項目: {s.get('要確認項目')}")

def run(input_path: str | Path = INPUT_PATH, output_file: str | Path | None = None) -> Path:
    summaries, details = check_folder(input_path)
    if not summaries:
        raise FileNotFoundError(f"対象PDFが見つかりません: {input_path}")
    out = Path(output_file) if output_file else get_timestamped_output_file()
    result_path = build_output(summaries, details, out)
    print_report(summaries)
    print(f"完了: {result_path}")
    return result_path

def clean_notebook_argv(argv: List[str]) -> List[str]:
    """Jupyter/VS Code Notebook の内部引数を除外する。"""
    cleaned = [argv[0]]
    skip_next = False
    for arg in argv[1:]:
        if skip_next:
            skip_next = False
            continue
        if arg.startswith("--f=") or arg.startswith("-f="):
            continue
        if arg in {"--f", "-f"}:
            skip_next = True
            continue
        if "kernel" in arg and arg.endswith(".json"):
            continue
        cleaned.append(arg)
    return cleaned

def main() -> None:
    parser = argparse.ArgumentParser(description="SPReAD 様式0/2/3/4 PDF セルフチェックツール")
    parser.add_argument("target", nargs="?", default=str(INPUT_PATH), help="検査対象の .pdf ファイル、またはフォルダ")
    parser.add_argument("--xlsx", dest="xlsx_path", help="結果Excelを保存するパス。未指定ならタイムスタンプ付きで保存")
    parser.add_argument("--json", dest="json_path", help="結果をJSONで保存するパス")
    parser.add_argument("--csv", dest="csv_path", help="詳細結果をCSVで保存するパス")

    argv = clean_notebook_argv(sys.argv)
    args, _unknown = parser.parse_known_args(argv[1:])

    summaries, details = check_folder(args.target)
    if not summaries:
        raise FileNotFoundError(f"対象PDFが見つかりません: {args.target}")

    xlsx_path = args.xlsx_path if args.xlsx_path else get_timestamped_output_file()
    build_output(summaries, details, xlsx_path)
    if args.json_path:
        write_json(summaries, details, args.json_path)
    if args.csv_path:
        write_csv(details, args.csv_path)
    print_report(summaries)
    print(f"完了: {xlsx_path}")

if __name__ == "__main__":
    main()
