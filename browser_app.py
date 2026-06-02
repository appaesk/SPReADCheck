from __future__ import annotations

import base64
import html
import json
import tempfile
from collections import Counter
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from form_self_check_v1 import run as run_form_checker
from research_plan_self_check_v1 import run as run_research_plan_checker, STATUS_LABEL_EN


ALLOWED_EXCEL_SUFFIXES = {".xlsx", ".xlsm"}
ALLOWED_PDF_SUFFIX = ".pdf"

# get all key and value from STATUS_LABEL_EN
STATUS_LABEL_ALL = set(STATUS_LABEL_EN.keys()) | set(STATUS_LABEL_EN.values())
STATUS_LABEL_NOT_OK = STATUS_LABEL_ALL - {"OK", "記入済み", "Completed", "金額OK", "Amount OK", "添付済み", "Attached"}


def format_preview_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    return str(value)


def workbook_preview(xlsx_path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(xlsx_path, data_only=True, read_only=True)
    previews: list[dict[str, object]] = []

    for worksheet in workbook.worksheets[:2]:
        rows = []
        max_rows = min(worksheet.max_row or 0, 50)
        max_cols = min(worksheet.max_column or 0, 10)
        for row in worksheet.iter_rows(min_row=1, max_row=max_rows, max_col=max_cols, values_only=True):
            rows.append([format_preview_value(cell) for cell in row])

        header = rows[0] if rows else []
        body = rows[1:] if len(rows) > 1 else []
        # alert_count = sum(1 for row in rows for cell in row if isinstance(cell, str) and "要確認" in cell)
        alert_count = sum(1 for row in rows for cell in row if isinstance(cell, str) and cell in STATUS_LABEL_NOT_OK)

        previews.append(
            {
                "title": worksheet.title,
                "header": header,
                "rows": body,
                "alert_count": alert_count,
                "row_count": max(0, len(rows) - 1),
                "column_count": len(header),
            }
        )

    return previews


def summarize_preview(previews: list[dict[str, object]]) -> dict[str, int]:
    status_words = Counter()
    for sheet in previews:
        for row in sheet["rows"]:
            for cell in row:
                # if isinstance(cell, str) and cell in {"OK", "要確認"}:
                if isinstance(cell, str) and cell in STATUS_LABEL_ALL:
                    status_words[cell] += 1

    return {
        "sheet_count": len(previews),
        "ok_count": status_words.get("OK", 0),
        # "needs_review_count": status_words.get("要確認", 0),
        "needs_review_count": sum(count for word, count in status_words.items() if word in STATUS_LABEL_NOT_OK),
    }


def render_preview_html(previews: list[dict[str, object]]) -> str:
    sections = []
    for sheet in previews:
        header_cells = "".join(f"<th>{html.escape(str(cell))}</th>" for cell in sheet["header"])
        body_rows = []
        for row in sheet["rows"]:
            # row_has_warning = any(isinstance(cell, str) and "要確認" in cell for cell in row)
            row_has_warning = any(isinstance(cell, str) and cell in STATUS_LABEL_NOT_OK for cell in row)
            row_class = ' class="row-warning"' if row_has_warning else ""
            body_cells = "".join(f"<td>{html.escape(str(cell))}</td>" for cell in row)
            body_rows.append(f"<tr{row_class}>{body_cells}</tr>")

        sections.append(
            """
            <section class="preview-card card">
              <div class="sheet-head">
                <div>
                  <h3>{title}</h3>
                  <p>Show only first 50 rows | {row_count} preview rows · {column_count} columns · {alert_count} review markers</p>
                </div>
              </div>
              <div class="table-wrap">
                <table>
                  <thead><tr>{header}</tr></thead>
                  <tbody>{body}</tbody>
                </table>
              </div>
            </section>
            """.format(
                title=html.escape(str(sheet["title"])),
                row_count=sheet["row_count"],
                column_count=sheet["column_count"],
                alert_count=sheet["alert_count"],
                header=header_cells,
                body="".join(body_rows),
            )
        )

    return "".join(sections)


def build_download_href(file_bytes: bytes) -> str:
    encoded = base64.b64encode(file_bytes).decode("ascii")
    return "data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64," + encoded


def run_check(mode: str, paths: list[str]) -> str:
    mode = (mode or "auto").lower()
    cleaned_paths = [Path(path) for path in paths]

    if not cleaned_paths:
        raise ValueError("Please choose at least one file.")

    suffixes = {path.suffix.lower() for path in cleaned_paths}
    if mode == "auto":
        if suffixes <= ALLOWED_EXCEL_SUFFIXES and len(cleaned_paths) == 1:
            mode = "research"
        elif suffixes == {ALLOWED_PDF_SUFFIX}:
            mode = "forms"
        else:
            raise ValueError("Auto mode needs one Excel file (.xlsx or .xlsm) or one or more PDF files.")

    with tempfile.TemporaryDirectory(prefix="spreadcheck_browser_") as temp_dir_name:
        temp_dir = Path(temp_dir_name)

        if mode == "research":
            if len(cleaned_paths) != 1:
                raise ValueError("Research Plan mode accepts one Excel file only.")
            if cleaned_paths[0].suffix.lower() not in ALLOWED_EXCEL_SUFFIXES:
                raise ValueError("Research Plan mode accepts .xlsx or .xlsm files only.")

            uploaded_path = temp_dir / cleaned_paths[0].name
            uploaded_path.write_bytes(cleaned_paths[0].read_bytes())
            output_path = temp_dir / "research_plan_result.xlsx"
            result_path = Path(run_research_plan_checker(uploaded_path, output_path))
            message = "Research Plan check complete."
        elif mode == "forms":
            if any(path.suffix.lower() != ALLOWED_PDF_SUFFIX for path in cleaned_paths):
                raise ValueError("Application Form mode accepts PDF files only.")

            input_dir = temp_dir / "forms"
            input_dir.mkdir(parents=True, exist_ok=True)
            for path in cleaned_paths:
                (input_dir / path.name).write_bytes(path.read_bytes())
            output_path = temp_dir / "form_result.xlsx"
            result_path = Path(run_form_checker(input_dir, output_path))
            message = "Application Form check complete."
        else:
            raise ValueError("Unsupported checker mode")

        result_bytes = result_path.read_bytes()
        previews = workbook_preview(result_path)
        summary = summarize_preview(previews)
        payload = {
            "sheet_count": summary["sheet_count"],
            "ok_count": summary["ok_count"],
            "needs_review_count": summary["needs_review_count"],
            "message": f"{message} The workbook is ready for download.",
            "preview_html": render_preview_html(previews),
            "download_href": build_download_href(result_bytes),
            "download_name": result_path.name,
        }
        return json.dumps(payload)