from __future__ import annotations

import io
import secrets
import tempfile
import time
from collections import Counter
from datetime import date, datetime
from pathlib import Path

from flask import Flask, abort, render_template, request, send_file, url_for
from openpyxl import load_workbook
from werkzeug.utils import secure_filename

from form_self_check_v1 import run as run_form_checker
from research_plan_self_check_v1 import run as run_research_plan_checker


BASE_DIR = Path(__file__).resolve().parent
UPLOAD_TTL_SECONDS = 30 * 60
MAX_STORED_RESULTS = 20
ALLOWED_EXCEL_SUFFIXES = {".xlsx", ".xlsm"}
ALLOWED_PDF_SUFFIX = ".pdf"

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 60 * 1024 * 1024

RESULT_STORE: dict[str, dict[str, object]] = {}


def cleanup_results() -> None:
    now = time.time()
    expired_tokens = [
        token
        for token, item in RESULT_STORE.items()
        if now - float(item["created_at"]) > UPLOAD_TTL_SECONDS
    ]
    for token in expired_tokens:
        RESULT_STORE.pop(token, None)

    if len(RESULT_STORE) <= MAX_STORED_RESULTS:
        return

    ordered_tokens = sorted(RESULT_STORE.items(), key=lambda pair: float(pair[1]["created_at"]))
    for token, _item in ordered_tokens[:-MAX_STORED_RESULTS]:
        RESULT_STORE.pop(token, None)


def normalize_mode(mode: str | None) -> str:
    if mode in {"research", "forms", "auto"}:
        return mode
    return "auto"


def allowed_suffix(path: str) -> str:
    return Path(path).suffix.lower()


def format_preview_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    return str(value)


def workbook_preview(xlsx_path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(xlsx_path, data_only=True, read_only=True)
    previews: list[dict[str, object]] = []

    for worksheet in workbook.worksheets[:2]:
        rows = []
        max_rows = min(worksheet.max_row or 0, 20)
        max_cols = min(worksheet.max_column or 0, 8)
        for row in worksheet.iter_rows(min_row=1, max_row=max_rows, max_col=max_cols, values_only=True):
            rows.append([format_preview_value(cell) for cell in row])

        header = rows[0] if rows else []
        body = rows[1:] if len(rows) > 1 else []
        alert_count = sum(1 for row in rows for cell in row if isinstance(cell, str) and "要確認" in cell)

        previews.append(
            {
                "title": worksheet.title,
                "header": header,
                "rows": body,
                "alert_count": alert_count,
                "row_count": max(0, len(rows) - 1),
                "column_count": len(header),
            }
        )

    return previews


def summarize_preview(previews: list[dict[str, object]]) -> dict[str, object]:
    status_words = Counter()
    for sheet in previews:
        for row in sheet["rows"]:
            for cell in row:
                if not isinstance(cell, str):
                    continue
                if cell in {"OK", "要確認"}:
                    status_words[cell] += 1

    return {
        "sheet_count": len(previews),
        "ok_count": status_words.get("OK", 0),
        "needs_review_count": status_words.get("要確認", 0),
    }


def store_result(filename: str, file_bytes: bytes, preview: list[dict[str, object]]) -> str:
    cleanup_results()
    token = secrets.token_urlsafe(24)
    RESULT_STORE[token] = {
        "filename": filename,
        "bytes": file_bytes,
        "preview": preview,
        "created_at": time.time(),
    }
    return token


def run_checker(mode: str, uploaded_files: list) -> tuple[str, bytes, list[dict[str, object]], dict[str, object]]:
    with tempfile.TemporaryDirectory(prefix="spreadcheck_") as temp_dir_name:
        temp_dir = Path(temp_dir_name)

        if mode == "research":
            upload = uploaded_files[0]
            uploaded_path = temp_dir / secure_filename(upload.filename or "research_plan.xlsx")
            upload.save(uploaded_path)
            output_path = temp_dir / "research_plan_result.xlsx"
            result_path = Path(run_research_plan_checker(uploaded_path, output_path))
        elif mode == "forms":
            input_dir = temp_dir / "forms"
            input_dir.mkdir(parents=True, exist_ok=True)
            for upload in uploaded_files:
                uploaded_path = input_dir / secure_filename(upload.filename or "form.pdf")
                upload.save(uploaded_path)
            output_path = temp_dir / "form_result.xlsx"
            result_path = Path(run_form_checker(input_dir, output_path))
        else:
            raise ValueError("Unsupported checker mode")

        result_bytes = result_path.read_bytes()
        preview = workbook_preview(result_path)
        summary = summarize_preview(preview)

        return result_path.name, result_bytes, preview, summary


@app.route("/", methods=["GET"])
def index() -> str:
    return render_template("index.html")


@app.route("/check", methods=["POST"])
def check() -> str:
    cleanup_results()
    mode = normalize_mode(request.form.get("mode"))
    uploads = [file for file in request.files.getlist("files") if file and file.filename]

    if not uploads:
        return render_template("index.html", error_message="Please choose at least one file.")

    suffixes = {allowed_suffix(file.filename) for file in uploads}
    if mode == "auto":
        if suffixes <= ALLOWED_EXCEL_SUFFIXES and len(uploads) == 1:
            mode = "research"
        elif suffixes == {ALLOWED_PDF_SUFFIX}:
            mode = "forms"
        else:
            return render_template(
                "index.html",
                error_message="Auto mode needs one Excel file (.xlsx or .xlsm) or one or more PDF files.",
            )

    if mode == "research":
        if len(uploads) != 1:
            return render_template("index.html", error_message="Research Plan mode accepts one Excel file only.")
        if allowed_suffix(uploads[0].filename or "") not in ALLOWED_EXCEL_SUFFIXES:
            return render_template(
                "index.html",
                error_message="Research Plan mode accepts .xlsx or .xlsm files only.",
            )
    elif mode == "forms":
        if any(allowed_suffix(file.filename or "") != ALLOWED_PDF_SUFFIX for file in uploads):
            return render_template(
                "index.html",
                error_message="Application Form mode accepts PDF files only.",
            )

    try:
        result_filename, result_bytes, preview, summary = run_checker(mode, uploads)
    except Exception as exc:  # noqa: BLE001 - surface checker failures to the user
        return render_template("index.html", error_message=str(exc))

    token = store_result(result_filename, result_bytes, preview)
    download_url = url_for("download_result", token=token)
    return render_template(
        "result.html",
        result_filename=result_filename,
        download_url=download_url,
        preview=preview,
        summary=summary,
        mode=mode,
    )


@app.route("/download/<token>", methods=["GET"])
def download_result(token: str):
    cleanup_results()
    result = RESULT_STORE.get(token)
    if not result:
        abort(404)

    buffer = io.BytesIO(result["bytes"])
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name=str(result["filename"]),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        max_age=0,
    )


@app.after_request
def prevent_cache(response):
    response.headers["Cache-Control"] = "no-store, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response


if __name__ == "__main__":
    app.run(debug=True)