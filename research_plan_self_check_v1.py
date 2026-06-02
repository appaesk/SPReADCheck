# -*- coding: utf-8 -*-
"""
SPReAD 様式1（研究計画調書）セルフチェックツール
（Microsoft Excel 専用・Jupyter Notebook／コマンドライン 対応）

本ツールは、第2回 SPReAD 公募の研究計画調書（様式1・Microsoft Excel）を確認し、
結果を Microsoft Excel ファイルとして出力します。

■ ライセンス・利用条件
本ツールは、SPReAD 第2回公募の応募者向けに配布するセルフチェック用ツールです。
応募者本人、応募者所属機関による応募書類の確認を目的とした利用に限ります。

・本ツールの商用利用は禁止します。
・ご自身の利用の範囲内での改変は可能ですが、改変したものを第三者へ配布することは禁止します。
・本ツールは、必ず公式配布元から入手してください。第三者により再配布されたものの利用は認められません。

本ツールの使用により生じた損害について、文部科学省は責任を負いません。

■ 対象ファイル
- 研究計画調書（様式1）の Microsoft Excel ファイル（.xlsx／.xlsm）のみを対象とします。
- PDF・Microsoft Word・画像ファイルは対象外です。
- 日本語版・英語版のいずれの様式にも対応しています。

■ 事前の準備
- 計算結果（文字数・合計金額）がファイルに保存された状態で実行してください。

■ 出力される Microsoft Excel
以下の2つのシートを作成します。
- 判定サマリー
- 詳細チェック

■ 必要なライブラリ
初回に一度、以下を実行してください。
    pip install openpyxl

■ 実行方法（コマンドライン）
    python research_plan_self_check.py

■ 実行方法（Jupyter Notebook）
    from research_plan_self_check import run
    run(r"C:\対象フォルダ")

■ 設定の変更
通常は、本ファイル上部の INPUT_PATH／OUTPUT_FILE／RECURSIVE のみを変更します。
これら以外はツール内部の判定仕様です。
"""

from pathlib import Path
from zipfile import ZipFile
from datetime import datetime, date
from numbers import Number
import posixpath
import re
import unicodedata
import warnings
import xml.etree.ElementTree as ET

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
import openpyxl.reader.excel as openpyxl_excel_reader


# ============================================================
# 設定（実行前に確認・変更する項目）
# ============================================================
# INPUT_PATH  : 確認対象。研究計画調書（Microsoft Excel）を含むフォルダ、または個々のファイルのパスを指定します。
#               既定値 Path(".") は、本ファイルを実行したフォルダを対象とします。
#               例) フォルダ指定 : Path(r"C:\Users\xxx\Desktop\応募書類")
#                   ファイル指定 : Path(r"C:\Users\xxx\Desktop\第2回_様式1_研究計画調書_1234_氏名.xlsx")
#               ※ Windows のパスは、先頭に r を付けて r"..." の形式で指定してください。
# OUTPUT_FILE : 結果ファイルの名称。実行のたびに末尾へ日時が付与されるため、前回の結果は上書きされません。
#               別の場所へ保存する場合は、フルパスで指定してください。
# RECURSIVE   : True にすると、指定フォルダ配下のサブフォルダも含めて確認します。
#               提出者ごとのフォルダをまとめた親フォルダを指定し、機関側で一括確認する際にご利用ください

INPUT_PATH = Path(".")
OUTPUT_FILE = Path("研究計画調書_セルフチェック結果.xlsx")
RECURSIVE = False

# ============================================================
# これより下は変更しないでください。

SUBMISSION_DATE_MIN = date(2026, 6, 2)
SUBMISSION_DATE_MAX = date(2026, 7, 3)


# openpyxl 読込時のエラー/重さ回避
def ignore_images_and_charts(archive, path):
    return [], []


openpyxl_excel_reader.find_images = ignore_images_and_charts
warnings.filterwarnings("ignore", message="Data Validation extension is not supported.*")


# 定数

# 入出力ファイルの拡張子と出力セル長の上限
ALLOWED_EXCEL_SUFFIXES = {".xlsx", ".xlsm"}
MAX_EXCEL_CELL_LENGTH = 32767
MAX_ZIP_MEMBER_BYTES = 20 * 1024 * 1024

# 計算結果（数式キャッシュ）が無いセルをツール側で算出したときに付ける注記。
RECALC_MARKER_JA = "（計算結果が未保存のため同じ計算式で本ツールが算出）"
RECALC_MARKER_EN = " (Excel value not stored; computed by the tool using the same formula)"
RECALC_SUMMARY_NOTE_JA = "※ 一部の文字数/語数セルの計算結果がファイルに保存されていなかったため、本ツールが同じ計算式で算出しました。表計算ソフトで開いて保存すると、ファイル内にも計算結果が保存されます。"
RECALC_SUMMARY_NOTE_EN = "* Some character/word-count cells were not yet calculated by Excel, so the tool computed them using the same formula (the result is valid). Opening and saving the file in Excel will also store the values in the file."

LANG_JA = "日本語"
LANG_EN = "English"
SHEET_KEYS = ["sheet1", "sheet2", "sheet3", "sheet4"]

EXPECTED_SHEETS_BY_LANGUAGE = {
    LANG_JA: {
        "sheet1": "研究計画調書_1枚目",
        "sheet2": "研究計画調書_2枚目",
        "sheet3": "研究計画調書_3枚目",
        "sheet4": "研究計画調書_4枚目",
    },
    LANG_EN: {
        "sheet1": "Research Plan_Sheet 1",
        "sheet2": "Research Plan_Sheet 2",
        "sheet3": "Research Plan_Sheet 3",
        "sheet4": "Research Plan_Sheet 4",
    },
}

LANGUAGE_CHECK_CELLS = {
    "sheet1": "M1",
    "sheet2": "E1",
    "sheet3": "N1",
    "sheet4": "G1",
}

ALLOWED_KUBUN_BY_LANGUAGE = {
    LANG_JA: {
        "大学",
        "高等専門学校",
        "公的研究機関",
        "民間企業",
        "非営利団体・公益法人",
        "その他（上記のいずれにも該当しない場合に限る）",
    },
    LANG_EN: {
        "University",
        "Technical College (Kosen)",
        "Public Research Institution",
        "Private Company",
        "Non-Profit Organization / Public Interest Corporation",
        "Other (only if none of the above apply)",
    },
}

ALLOWED_APPLICANT_ATTRIBUTES_BY_LANGUAGE = {
    LANG_JA: {
        "教員",
        "研究員(ポスドク含む)",
        "博士課程学生",
        "修士課程学生",
        "学部学生",
        "技術職員・事務職員・URA等",
        "技術者・開発者",
        "その他（上記のいずれにも該当しない場合に限る）",
    },
    LANG_EN: {
        "Faculty Member",
        "Researcher (including Postdoc)",
        "Doctoral Student",
        "Master's Student",
        "Undergraduate Student",
        "Technical / Administrative Staff / URA, etc.",
        "Engineer / Developer",
        "Other (only if none of the above apply)",
    },
}

ALLOWED_RESEARCH_AREAS_BY_LANGUAGE = {
    LANG_JA: {
        "臨床科学",
        "生命科学・薬学",
        "化学",
        "機械・社会基盤・エネルギー工学",
        "材料・プロセス・応用医工学",
        "電気工学・電子工学・情報科学・コンピューターサイエンス",
        "数学・物理学・地球科学",
        "農学・環境学・生態学",
        "社会科学",
        "芸術・人文科学",
    },
    LANG_EN: {
        "Clinical Science",
        "Life Sciences / Pharmacy",
        "Chemistry",
        "Mechanical / Infrastructure / Energy Engineering",
        "Materials / Process / Applied Biomedical Engineering",
        "Electrical Engineering / Electronic Engineering / Information Science / Computer Science",
        "Mathematics / Physics / Earth Sciences",
        "Agriculture / Environmental Science / Ecology",
        "Social Sciences",
        "Arts / Humanities",
    },
}

ALLOWED_MAIN_USE_CASES_BY_LANGUAGE = {
    LANG_JA: {
        "1.学習用データセット構築",
        "2.既存モデルの適応",
        "3.AIモデル開発",
        "4.既存モデル評価",
        "5.実験自動化・自律化",
        "6.シミュレーション・デジタルツイン",
        "7.発見・設計支援",
        "8.高度データ解析・モデリング",
        "9.その他",
    },
    LANG_EN: {
        "1. Training Dataset Construction",
        "2. Adaptation of Existing Models",
        "3. AI Model Development",
        "4. Evaluation of Existing Models",
        "5. Experimental Automation / Autonomization",
        "6. Simulation / Digital Twin",
        "7. Discovery / Design Support",
        "8. Advanced Data Analysis / Modeling",
        "9. Other",
    },
}

# 研究計画調書_1枚目: 選択セル（Y を入れる）と、対応する選択肢ラベルのセル
SUB_USE_CASE_CELLS = ["C32", "E32", "I32", "G32", "C33", "G33", "E33", "I33"]
AI_USAGE_CELLS = ["C39", "E39", "G39", "I39", "K39", "C40", "E40", "G40", "I40", "K40"]

SUB_USE_CASE_LABEL_MAP = {
    "C32": "D32",
    "E32": "F32",
    "I32": "J32",
    "G32": "H32",
    "C33": "D33",
    "G33": "H33",
    "E33": "F33",
    "I33": "J33",
}

AI_USAGE_LABEL_MAP = {
    "C39": "D39",
    "E39": "F39",
    "G39": "H39",
    "I39": "J39",
    "K39": "L39",
    "C40": "D40",
    "E40": "F40",
    "G40": "H40",
    "I40": "J40",
    "K40": "L40",
}

# 研究計画調書_2枚目: 各項目の入力行と、文字数（日本語）／語数（英語）の下限・上限
TEXT_COUNT_RULES_BY_LANGUAGE = {
    LANG_JA: {
        "研究目的": {"row": 8, "min": 80, "max": 400, "optional": False},
        "研究方法": {"row": 9, "min": 160, "max": 800, "optional": False},
        "AI利活用の妥当性・実現可能性": {"row": 10, "min": 160, "max": 800, "optional": False},
        "達成目標": {"row": 11, "min": 100, "max": 500, "optional": False},
        "AI利活用のノウハウ抽出・共有の実現計画": {"row": 12, "min": 60, "max": 300, "optional": False},
        "成果の公開方針（任意）": {"row": 13, "min": None, "max": 150, "optional": True},
    },
    LANG_EN: {
        "研究目的": {"row": 8, "min": 48, "max": 240, "optional": False},
        "研究方法": {"row": 9, "min": 96, "max": 480, "optional": False},
        "AI利活用の妥当性・実現可能性": {"row": 10, "min": 96, "max": 480, "optional": False},
        "達成目標": {"row": 11, "min": 60, "max": 300, "optional": False},
        "AI利活用のノウハウ抽出・共有の実現計画": {"row": 12, "min": 36, "max": 180, "optional": False},
        "成果の公開方針（任意）": {"row": 13, "min": None, "max": 90, "optional": True},
    },
}

# 研究計画調書_3枚目・4枚目: 経費明細の範囲
DETAIL_RANGES = {
    "設備備品費の明細": "D11:J30",
    "消耗品費の明細": "M11:N30",
    "謝金の明細": "D40:E59",
    "国内旅費の明細": "H40:J59",
    "外国旅費の明細": "M40:N59",
    "その他の明細": "D69:E88",
    "費用詳細 API費用": "D9:F18",
    "費用詳細 計算資源費用 (クラウドGPU含む)": "D22:G31",
}


# 共通ヘルパー

def normalize_text(value):
    if value is None:
        return ""
    text = unicodedata.normalize("NFKC", str(value))
    text = text.replace("\r", "").replace("\n", "")
    text = text.replace("\u3000", " ").replace("\xa0", " ")
    return text.strip()


def normalize_sheet_name_loose(sheet_name):
    text = unicodedata.normalize("NFKC", str(sheet_name))
    text = text.replace(" ", "").replace("　", "")
    return text.strip()


def normalize_file_name_for_check(name):
    text = unicodedata.normalize("NFKC", str(name))
    text = text.replace(" ", "").replace("　", "")
    return text.lower()

def get_internal_sheet_name(sheet_key):
    return EXPECTED_SHEETS_BY_LANGUAGE[LANG_JA][sheet_key]


def get_expected_sheet_name(language, sheet_key):
    return EXPECTED_SHEETS_BY_LANGUAGE.get(language, EXPECTED_SHEETS_BY_LANGUAGE[LANG_JA])[sheet_key]


def detect_workbook_language(wb):
    """1〜4枚目の指定セルにある 日本語 / English から版を判定する。"""
    for language in (LANG_JA, LANG_EN):
        values = []
        all_sheets_found = True
        for sheet_key in SHEET_KEYS:
            expected = get_expected_sheet_name(language, sheet_key)
            actual = find_sheet_name_for_transfer(wb, expected)
            if not actual:
                all_sheets_found = False
                break
            values.append(normalize_text(get_cell_value(wb[actual], LANGUAGE_CHECK_CELLS[sheet_key])))
        if all_sheets_found and all(value == language for value in values):
            return language

    # フォールバック: シート名がそろっている版を優先する。
    for language in (LANG_JA, LANG_EN):
        if all(find_sheet_name_for_transfer(wb, get_expected_sheet_name(language, key)) for key in SHEET_KEYS):
            return language
    return LANG_JA


def get_allowed_kubun(language):
    return ALLOWED_KUBUN_BY_LANGUAGE.get(language, ALLOWED_KUBUN_BY_LANGUAGE[LANG_JA])


def get_allowed_applicant_attributes(language):
    return ALLOWED_APPLICANT_ATTRIBUTES_BY_LANGUAGE.get(language, ALLOWED_APPLICANT_ATTRIBUTES_BY_LANGUAGE[LANG_JA])


def get_allowed_research_areas(language):
    return ALLOWED_RESEARCH_AREAS_BY_LANGUAGE.get(language, ALLOWED_RESEARCH_AREAS_BY_LANGUAGE[LANG_JA])


def get_allowed_main_use_cases(language):
    return ALLOWED_MAIN_USE_CASES_BY_LANGUAGE.get(language, ALLOWED_MAIN_USE_CASES_BY_LANGUAGE[LANG_JA])


def is_main_use_case_other(value, language):
    cleaned = clean_choice(value)
    if language == LANG_EN:
        return cleaned in {"9.Other", "Other"} or cleaned.endswith("Other")
    return cleaned in {"9.その他", "その他"} or cleaned.endswith("その他")


def safe_text_for_output(value):
    """Excel出力用に外部入力文字列を安全化する。

    - HTML/スクリプト断片として解釈され得る < > をエスケープする
    - Excel数式インジェクションにつながる先頭文字を無害化する
    - 制御文字とExcelセル上限超過を抑制する
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return value
    if isinstance(value, Number):
        return int(value) if isinstance(value, float) and value.is_integer() else value
    if isinstance(value, (datetime, date)):
        return value

    text = str(value)
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]", "", text)
    text = text.replace("<", "&lt;").replace(">", "&gt;")

    formula_probe = text.lstrip(" \t\n")
    if formula_probe.startswith(("=", "+", "@")) or re.match(r"^-[A-Za-z0-9_().]", formula_probe):
        text = "'" + text

    return text[:MAX_EXCEL_CELL_LENGTH]


def safe_display_text(value):
    safe_value = safe_text_for_output(value)
    if isinstance(safe_value, (datetime, date)):
        return safe_value.isoformat()
    return str(safe_value)


def safe_resolve_path(path):
    """パストラバーサル検知に使う正規化済み絶対パスを返す。"""
    path_obj = Path(path).expanduser()
    return path_obj.resolve(strict=False)


def is_path_within_directory(child_path, parent_dir):
    child = safe_resolve_path(child_path)
    parent = safe_resolve_path(parent_dir)
    try:
        child.relative_to(parent)
        return True
    except ValueError:
        return False


def get_configured_input_base_dir():
    configured_input = safe_resolve_path(INPUT_PATH)
    if configured_input.suffix.lower() in ALLOWED_EXCEL_SUFFIXES:
        return configured_input.parent
    return configured_input


def validate_excel_file_path(file_path, base_dir=None, must_exist=True):
    """処理対象Excelを拡張子・所在ディレクトリで許可リスト検証する。"""
    candidate = safe_resolve_path(file_path)
    if candidate.name.startswith("~$"):
        raise ValueError("一時ファイルは処理対象外です")
    if candidate.suffix.lower() not in ALLOWED_EXCEL_SUFFIXES:
        raise ValueError("処理対象は .xlsx または .xlsm のみです")
    if must_exist and not candidate.is_file():
        raise FileNotFoundError(f"対象Excelが見つかりません: {candidate}")
    if base_dir is not None and not is_path_within_directory(candidate, base_dir):
        raise ValueError("許可された入力フォルダ外のファイルは処理できません")
    return candidate


def validate_output_file_path(output_file):
    """出力先を .xlsx に限定し、親ディレクトリを正規化する。"""
    raw_path = Path(output_file).expanduser()
    if any(part == ".." for part in raw_path.parts):
        raise ValueError("出力ファイルパスに '..' は使用できません")
    candidate = safe_resolve_path(raw_path)
    if candidate.suffix.lower() != ".xlsx":
        raise ValueError("出力ファイルの拡張子は .xlsx にしてください")
    if candidate.name.startswith("~$"):
        raise ValueError("出力ファイル名が不正です")
    return candidate


def get_timestamped_output_file(output_language=LANG_JA):
    """出力ファイル名を言語別にし、末尾に出力時間を付ける。"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir = safe_resolve_path(Path(OUTPUT_FILE).expanduser().parent)
    if output_language == LANG_EN:
        filename = f"Research_Plan_Self_Check_Result_{timestamp}.xlsx"
    else:
        filename = f"研究計画調書_セルフチェック結果_{timestamp}.xlsx"
    return validate_output_file_path(output_dir / filename)


def normalize_zip_member_path(member_path):
    """OOXML ZIP内部パスを安全な相対パスに正規化する。"""
    text = str(member_path).replace("\\", "/")
    if "\x00" in text:
        raise ValueError("ZIP内部パスに不正な文字が含まれています")
    normalized = posixpath.normpath(text)
    if normalized in {"", ".", ".."} or normalized.startswith("../") or normalized.startswith("/"):
        raise ValueError("ZIP内部パスが不正です")
    return normalized


def safe_zip_read(zip_file, member_path, names=None):
    member = normalize_zip_member_path(member_path)
    names = names if names is not None else set(zip_file.namelist())
    if member not in names:
        raise KeyError(member)
    info = zip_file.getinfo(member)
    if info.file_size > MAX_ZIP_MEMBER_BYTES:
        raise ValueError("ZIP内部ファイルが大きすぎます")
    return zip_file.read(member)


def safe_ooxml_target(base_part_path, target, required_prefix="xl/"):
    """OOXML relationship のTargetを、期待するZIP内部領域に限定して解決する。"""
    target_text = str(target).replace("\\", "/")
    if target_text.startswith("/"):
        resolved = normalize_zip_member_path(target_text.lstrip("/"))
    else:
        base_part = normalize_zip_member_path(base_part_path)
        resolved = normalize_zip_member_path(posixpath.join(posixpath.dirname(base_part), target_text))
    if required_prefix and not resolved.startswith(required_prefix):
        raise ValueError("OOXML参照先が許可された領域外です")
    return resolved


def judge_file_name(file_path, language=LANG_JA):
    """指定された研究計画調書ファイル名かを確認する。"""
    if file_path.name.startswith("~$"):
        return "NG"
    if file_path.suffix.lower() not in ALLOWED_EXCEL_SUFFIXES:
        return "NG"
    stem = normalize_file_name_for_check(file_path.stem)

    if language == LANG_EN:
        # 2nd_Form1_Research Plan_e-Rad Institution code_Name
        pattern = r"^2nd_form1_researchplan_([0-9]+)_([^_]+)$"
        m = re.fullmatch(pattern, stem)
        if not m:
            return "NG"
        name_part = m.group(2)
        if not re.search(r"[a-z]", name_part):
            return "NG"
        return "OK"

    # 第2回_様式1_研究計画調書_e-Rad所属機関コード_氏名
    pattern = r"^第2回_様式1_研究計画調書_([0-9]+)_([^_]+)$"
    m = re.fullmatch(pattern, stem)
    if not m:
        return "NG"
    if not m.group(2):
        return "NG"
    return "OK"


def find_sheet_name_for_transfer(wb, expected_sheet_name):
    """期待シート名を探す。半角/全角スペースの差分のみ許容する。"""
    if expected_sheet_name in wb.sheetnames:
        return expected_sheet_name
    expected_loose = normalize_sheet_name_loose(expected_sheet_name)
    for actual in wb.sheetnames:
        if normalize_sheet_name_loose(actual) == expected_loose:
            return actual
    return ""


def is_blank(value):
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def has_input(value):
    if value is None:
        return False
    if isinstance(value, bool):
        return True
    if isinstance(value, Number):
        return True
    if isinstance(value, (datetime, date)):
        return True
    return normalize_text(value) != ""


def get_cell_value(ws, cell_addr):
    value = ws[cell_addr].value
    if value is not None:
        return value
    for merged_range in ws.merged_cells.ranges:
        if cell_addr in merged_range:
            value = ws.cell(merged_range.min_row, merged_range.min_col).value
            return "" if value is None else value
    return ""


def to_number(value):
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, Number):
        return int(value) if isinstance(value, float) and value.is_integer() else value
    text = unicodedata.normalize("NFKC", str(value)).strip()
    text = text.replace(",", "").replace("円", "").replace("¥", "").replace("￥", "")
    text = text.replace(" ", "").replace("　", "")
    if text == "":
        return None
    try:
        num = float(text)
        return int(num) if num.is_integer() else num
    except Exception:
        return None


def to_digits_text(value):
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y%m%d")
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = unicodedata.normalize("NFKC", str(value)).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return re.sub(r"\D", "", text)


def clean_choice(value):
    text = normalize_text(value)
    text = text.replace(" ", "").replace("　", "")
    return text


def is_allowed_choice(value, allowed_set):
    text = clean_choice(value)
    allowed_clean = {clean_choice(x) for x in allowed_set}
    return text in allowed_clean


def is_y_or_blank(value):
    return not has_input(value) or normalize_text(value).upper() == "Y"


def y_selection_status(values):
    """選択セルは空欄またはYのみ許容。

    - 1つ以上Yがあり、その他が空欄またはY: 記入済み
    - 全セル空欄: 未記入
    - Y・空欄以外が1つでもある: 表記ゆれ
    """
    if any(not is_y_or_blank(v) for v in values):
        return "表記ゆれ"
    if any(normalize_text(v).upper() == "Y" for v in values):
        return "記入済み"
    return "未記入"


def is_y(value):
    return normalize_text(value).upper() == "Y"


def is_student_attribute(value, language=None):
    japanese_students = {"博士課程学生", "修士課程学生", "学部学生"}
    english_students = {"DoctoralStudent", "Master'sStudent", "UndergraduateStudent"}
    cleaned = clean_choice(value)
    if language == LANG_EN:
        return cleaned in english_students
    if language == LANG_JA:
        return cleaned in japanese_students
    return cleaned in japanese_students or cleaned in english_students


def judge_ymd_parts(year_value, month_value, day_value):
    if not (has_input(year_value) and has_input(month_value) and has_input(day_value)):
        return "未記入"
    year_digits = to_digits_text(year_value)
    month_digits = to_digits_text(month_value)
    day_digits = to_digits_text(day_value)
    if len(year_digits) == 4 and 1 <= len(month_digits) <= 2 and 1 <= len(day_digits) <= 2:
        return "記入済み"
    return "桁数間違い"


def judge_submission_date(year_value, month_value, day_value):
    """提出日が指定期間内か判定する。

    - 未入力・桁数不正は従来どおりの判定を返す
    - 実在しない日付、または 2026/6/2〜2026/7/3 の範囲外は「日付要確認」
    """
    base_status = judge_ymd_parts(year_value, month_value, day_value)
    if base_status != "記入済み":
        return base_status

    year_digits = to_digits_text(year_value)
    month_digits = to_digits_text(month_value)
    day_digits = to_digits_text(day_value)
    try:
        submitted_date = date(int(year_digits), int(month_digits), int(day_digits))
    except ValueError:
        return "日付要確認"

    if SUBMISSION_DATE_MIN <= submitted_date <= SUBMISSION_DATE_MAX:
        return "記入済み"
    return "日付要確認"


def judge_birthdate(year_value, month_value, day_value):
    return judge_ymd_parts(year_value, month_value, day_value)


def judge_text_count_by_e_only(count_value, min_chars, max_chars, optional=False):
    """E列の文字数カウント値のみを見て判定する。F列は参照しない。"""
    num = to_number(count_value)
    num = 0 if num is None else num

    if optional and num == 0:
        return "未記入"
    if min_chars is not None and num < min_chars:
        return "文字数不足"
    if max_chars is not None and num > max_chars:
        return "文字数超過"
    return "記入済み"

def has_detail_row_input(value):
    """費目明細の行カウント用。

    テンプレート側で自動計算列に 0 が入っている行を、入力済み行として数えない。
    そのため、空欄・0・0.0・文字列の0は未入力扱いにする。
    """
    if not has_input(value):
        return False
    num = to_number(value)
    if num == 0:
        return False
    return True


def analyze_detail_range_rows(ws, range_addr):
    """明細範囲を行単位で判定する。

    - 1行の全列が空欄/0なら未記入行
    - 1行のどこか1セルでも入力がある場合、その行の範囲内全列が入力必須
    - 入力行で1列でも空欄/0があれば記入不足
    """
    min_col, min_row, max_col, max_row = range_boundaries(range_addr)
    complete_count = 0
    incomplete_rows = []
    for row in range(min_row, max_row + 1):
        row_values = [get_cell_value(ws, f"{get_column_letter(col)}{row}") for col in range(min_col, max_col + 1)]
        entered_flags = [has_detail_row_input(v) for v in row_values]
        if not any(entered_flags):
            continue
        if all(entered_flags):
            complete_count += 1
        else:
            incomplete_rows.append(row)
    if incomplete_rows:
        status = "記入不足"
        value = f"{complete_count}項目記入済み / 記入不足行: {', '.join(map(str, incomplete_rows))}"
    elif complete_count > 0:
        status = f"{complete_count}項目記入済み"
        value = complete_count
    else:
        status = "未記入"
        value = 0
    return status, value, complete_count, incomplete_rows


def count_filled_rows(ws, range_addr):
    status, value, complete_count, incomplete_rows = analyze_detail_range_rows(ws, range_addr)
    return complete_count


def count_result(count):
    return f"{count}項目記入済み" if count > 0 else "未記入"


def join_nonblank_values(values, separator=" / "):
    """出力用に、空欄・0を除いた値を指定順で連結する。"""
    output_values = []
    for value in values:
        if has_detail_row_input(value):
            output_values.append(str(value))
    return separator.join(output_values)


def selected_right_labels(get, selection_cells, label_map):
    """Y が入っている選択セルについて、右隣セルのラベルを指定順で返す。"""
    labels = []
    for cell in selection_cells:
        if is_y(get(cell)):
            label = get(label_map[cell])
            if has_input(label):
                labels.append(str(label))
    return " / ".join(labels)


def display_values_from_cells(get, cells):
    return join_nonblank_values([get(cell) for cell in cells])


OK_UNWRITTEN_ITEMS = {
    "成果の公開方針（任意）",
    "研究業績等",
    "画像の添付",
    "設備備品費の明細",
    "消耗品費の明細",
    "謝金の明細",
    "国内旅費の明細",
    "外国旅費の明細",
    "その他の明細",
    "費用詳細 API費用",
    "費用詳細 計算資源費用 (クラウドGPU含む)",
}


def is_ok_detail_status(detail):
    status = detail.get("判定", "")
    item = detail.get("項目", "")
    if status in {"OK", "記入済み", "添付済み", "対象外", "金額OK"}:
        return True
    if isinstance(status, str) and status.endswith("項目記入済み"):
        return True
    if status in {"未記入", "添付なし"} and item in OK_UNWRITTEN_ITEMS:
        return True
    return False


# 数式の計算結果の補完

def get_formula_aware_value(wb_values, wb_formula, sheet_name, cell_addr):
    actual = find_sheet_name_for_transfer(wb_values, sheet_name)
    if not actual:
        return ""
    value = get_cell_value(wb_values[actual], cell_addr)
    if not is_blank(value):
        return value
    if wb_formula is None:
        return ""
    actual_f = find_sheet_name_for_transfer(wb_formula, sheet_name)
    if not actual_f:
        return ""
    raw = get_cell_value(wb_formula[actual_f], cell_addr)
    if not isinstance(raw, str) or not raw.startswith("="):
        return "" if raw is None else raw
    # SUM(単純範囲) のみ補完。その他は空欄扱い。
    m = re.fullmatch(r"=SUM\(([^)]+)\)", raw.strip(), flags=re.I)
    if not m:
        return ""
    total = 0
    found = False
    for part in re.split(r"[,;]", m.group(1)):
        part = part.replace("$", "").strip()
        if "!" in part:
            sheet_part, ref = part.rsplit("!", 1)
            ref_sheet = sheet_part.strip("'")
        else:
            ref_sheet, ref = sheet_name, part
        ref_actual = find_sheet_name_for_transfer(wb_values, ref_sheet)
        if not ref_actual:
            continue
        try:
            min_col, min_row, max_col, max_row = range_boundaries(ref)
        except Exception:
            continue
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                n = to_number(get_cell_value(wb_values[ref_actual], f"{get_column_letter(c)}{r}"))
                if n is not None:
                    total += n
                    found = True
    return total if found else ""


def count_chars_excel_len(text):
    """日本語版テンプレートの =LEN($D) と同じ文字数を返す。

    空白・改行・記号も含め、全ての文字を 1 文字として数える（Excel の LEN と一致）。
    """
    return len(str(text))


def count_words_excel(text):
    """英語版テンプレートの語数カウント式と同じ語数を返す。

    元の式: =IF($D="",0,LEN(TRIM($D))-LEN(SUBSTITUTE(TRIM($D)," ",""))+1)
    すなわち、半角スペースのみを区切りとみなし（Excel の TRIM 準拠で前後を除去し
    連続スペースを 1 つに圧縮）、語数 = スペース数 + 1 とする。
    """
    s = str(text)
    if s == "":
        return 0
    trimmed = re.sub(r" {2,}", " ", s.strip(" "))
    return trimmed.count(" ") + 1


def recompute_count_for_sheet2(body_value, language):
    """E 列（文字数/語数）が Excel で未計算のとき、本文 D 列からテンプレートと
    同じ計算式で文字数（日本語）／語数（英語）を再計算して返す。

    本文が空なら 0 を返す（= 未記入相当）。LibreOffice での実測により、本文に
    対するこの再計算値はテンプレートの数式結果と一致することを確認済み。
    """
    if not has_input(body_value):
        return 0
    if language == LANG_EN:
        return count_words_excel(body_value)
    return count_chars_excel_len(body_value)


# 画像カウント

def local_name(tag):
    return tag.split("}", 1)[1] if "}" in tag else tag


def get_sheet_xml_path(xlsx_path, sheet_name):
    rel_attr = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
    safe_xlsx_path = validate_excel_file_path(xlsx_path, get_configured_input_base_dir(), must_exist=True)
    with ZipFile(safe_xlsx_path) as z:
        names = set(z.namelist())
        root = ET.fromstring(safe_zip_read(z, "xl/workbook.xml", names))
        target_rid = ""
        for elem in root.iter():
            if local_name(elem.tag) == "sheet" and elem.attrib.get("name") == sheet_name:
                target_rid = elem.attrib.get(rel_attr, "")
                break
        if not target_rid:
            return ""
        rels_root = ET.fromstring(safe_zip_read(z, "xl/_rels/workbook.xml.rels", names))
        for rel in rels_root:
            if rel.attrib.get("Id") == target_rid:
                target = rel.attrib.get("Target", "")
                try:
                    if target.startswith("/"):
                        candidate = normalize_zip_member_path(target.lstrip("/"))
                    else:
                        candidate = normalize_zip_member_path(posixpath.join("xl", target))
                    if not candidate.startswith("xl/"):
                        return ""
                    return candidate
                except ValueError:
                    return ""
    return ""


def get_rels_path_for_part(part_path):
    folder = posixpath.dirname(part_path)
    filename = posixpath.basename(part_path)
    return posixpath.join(folder, "_rels", filename + ".rels")


def resolve_part_target(base_part_path, target):
    return safe_ooxml_target(base_part_path, target, required_prefix="xl/")


def count_images_on_sheet_from_xml(xlsx_path, sheet_name):
    try:
        safe_xlsx_path = validate_excel_file_path(xlsx_path, get_configured_input_base_dir(), must_exist=True)
        sheet_xml_path = get_sheet_xml_path(safe_xlsx_path, sheet_name)
        if not sheet_xml_path:
            return 0
        image_count = 0
        with ZipFile(safe_xlsx_path) as z:
            names = set(z.namelist())
            rels_path = get_rels_path_for_part(sheet_xml_path)
            if rels_path in names:
                rels_root = ET.fromstring(safe_zip_read(z, rels_path, names))
                drawing_paths = []
                for rel in rels_root:
                    if rel.attrib.get("Type", "").endswith("/drawing"):
                        try:
                            drawing_paths.append(resolve_part_target(sheet_xml_path, rel.attrib.get("Target", "")))
                        except ValueError:
                            continue
                for drawing_path in drawing_paths:
                    if drawing_path not in names:
                        continue
                    drawing_root = ET.fromstring(safe_zip_read(z, drawing_path, names))
                    pics = sum(1 for elem in drawing_root.iter() if local_name(elem.tag) == "pic")
                    blips = sum(1 for elem in drawing_root.iter() if local_name(elem.tag) == "blip")
                    image_count += pics if pics else blips
            # IMAGE関数 / DISPIMG関数も画像ありとしてカウント
            if sheet_xml_path in names:
                sheet_root = ET.fromstring(safe_zip_read(z, sheet_xml_path, names))
                for cell in sheet_root.iter():
                    if local_name(cell.tag) != "c":
                        continue
                    for child in cell:
                        if local_name(child.tag) == "f" and child.text:
                            formula = child.text.upper()
                            if "IMAGE(" in formula or "DISPIMG(" in formula:
                                image_count += 1
                                break
        return image_count
    except Exception:
        return 0


# チェック本体

MAJOR_ITEM_MAP = {
    "ファイル名": "-",
    "タブ名": "-",

    "提出日": "応募者情報",
    "e-Rad 研究者番号": "応募者情報",
    "メールアドレス": "応募者情報",
    "研究代表者氏名": "応募者情報",
    "生年月日": "応募者情報",
    "e-Rad所属機関コード": "応募者情報",
    "所属機関": "応募者情報",
    "部局": "応募者情報",
    "職": "応募者情報",
    "所属機関の区分": "応募者情報",
    "応募者属性の区分": "応募者情報",
    "学生フラグ": "応募者情報",
    "研究領域": "研究課題情報",
    "メインユースケース分類": "研究課題情報",
    "メインユースケース分類（その他）": "研究課題情報",
    "サブユースケース選択": "研究課題情報",
    "研究課題名": "研究課題情報",
    "AI活用度選択": "AI活用度",
    "現在の具体的な活用方法": "AI活用度",
    "研究経費": "研究経費",

    "研究目的": "研究内容",
    "研究方法": "研究内容",
    "AI利活用の妥当性・実現可能性": "研究内容",
    "達成目標": "研究内容",
    "AI利活用のノウハウ抽出・共有の実現計画": "研究内容",
    "成果の公開方針（任意）": "研究内容",
    "研究業績等": "研究内容",
    "画像の添付": "研究内容",

    "設備備品費の明細": "1.設備備品費、消耗品費",
    "消耗品費の明細": "1.設備備品費、消耗品費",
    "設備備品費、消耗品費の必要性": "1.設備備品費、消耗品費",
    "謝金の明細": "2.謝金、旅費",
    "国内旅費の明細": "2.謝金、旅費",
    "外国旅費の明細": "2.謝金、旅費",
    "謝金、旅費の必要性": "2.謝金、旅費",
    "その他の明細": "3.その他費用",
    "その他費用の必要性": "3.その他費用",

    "費用詳細 API費用": "API/GPU費用詳細",
    "費用詳細 計算資源費用 (クラウドGPU含む)": "API/GPU費用詳細",
}

ITEM_LABEL_EN = {
    "ファイル名": "File Name",
    "タブ名": "Sheet Name",
    "提出日": "Submission Date",
    "e-Rad 研究者番号": "e-Rad Researcher ID",
    "メールアドレス": "Email Address",
    "研究代表者氏名": "Principal Investigator Name",
    "生年月日": "Date of Birth",
    "e-Rad所属機関コード": "e-Rad Institution Code",
    "所属機関": "Institution",
    "部局": "Department / Division",
    "職": "Position / Title",
    "所属機関の区分": "Institution Category",
    "応募者属性の区分": "Applicant Attribute Category",
    "学生フラグ": "Student Flag",
    "研究領域": "Research Area",
    "メインユースケース分類": "Main Use Case Classification",
    "メインユースケース分類（その他）": "Main Use Case Classification (Other)",
    "サブユースケース選択": "Sub Use Case Selection",
    "研究課題名": "Research Project Title",
    "AI活用度選択": "AI Utilization Level Selection",
    "現在の具体的な活用方法": "Specific Current Usage Methods",
    "研究経費": "Research Expenses",
    "研究目的": "Research Objectives",
    "研究方法": "Research Methods",
    "AI利活用の妥当性・実現可能性": "Rationale and Feasibility of AI Utilization",
    "達成目標": "Achievement Goals",
    "AI利活用のノウハウ抽出・共有の実現計画": "Plan for Extracting and Sharing AI Utilization Know-How",
    "成果の公開方針（任意）": "Publication Policy for Research Outcomes",
    "研究業績等": "Research Achievements, etc.",
    "画像の添付": "Image Attachment",
    "設備備品費の明細": "Equipment and Fixtures Breakdown",
    "消耗品費の明細": "Consumable Supplies Breakdown",
    "設備備品費、消耗品費の必要性": "Necessity of Equipment, Fixtures and Consumable Supplies",
    "謝金の明細": "Honoraria Breakdown",
    "国内旅費の明細": "Domestic Travel Expense Breakdown",
    "外国旅費の明細": "International Travel Expense Breakdown",
    "謝金、旅費の必要性": "Necessity of Honoraria and Travel Expenses",
    "その他の明細": "Other Expenses Breakdown",
    "その他費用の必要性": "Necessity of Other Expenses",
    "費用詳細 API費用": "Cost Details: API Costs",
    "費用詳細 計算資源費用 (クラウドGPU含む)": "Cost Details: Computational Resource Costs (including Cloud GPU)",
}

CATEGORY_LABEL_EN = {
    "-": "",
    "応募者情報": "Applicant Information",
    "研究課題情報": "Research Project Information",
    "AI活用度": "AI Utilization Level",
    "研究経費": "Research Expenses",
    "研究内容": "Research Content",
    "1.設備備品費、消耗品費": "1. Equipment, Fixtures and Consumable Supplies",
    "2.謝金、旅費": "2. Honoraria and Travel Expenses",
    "3.その他費用": "3. Other Expenses",
    "API/GPU費用詳細": "API/GPU Expense Details",
}

SHEET_LABEL_EN = {
    "研究計画調書_1枚目": "Research Plan_Sheet 1",
    "研究計画調書_2枚目": "Research Plan_Sheet 2",
    "研究計画調書_3枚目": "Research Plan_Sheet 3",
    "研究計画調書_4枚目": "Research Plan_Sheet 4",
    "-": "-",
}

SUMMARY_HEADER_EN = {
    "ファイル名": "File Name",
    "ファイルパス": "File Path",
    "総合判定": "Overall Result",
    "指定されたファイル名になっているか": "Whether the file name follows the specified format",
    "タブ名を変更していないか": "Whether sheet names have not been changed",
    "研究計画調書_1枚目": "Research Plan_Sheet 1",
    "研究計画調書_2枚目": "Research Plan_Sheet 2",
    "研究計画調書_3枚目": "Research Plan_Sheet 3",
    "研究計画調書_4枚目": "Research Plan_Sheet 4",
    "要確認項目": "Items Needing Review",
    "エラー内容": "Error Details",
}

DETAIL_HEADER_EN = {
    "ファイル名": "File Name",
    "タブ名": "Sheet Name",
    "大項目": "Category",
    "項目": "Item",
    "判定": "Result",
    "参照セル": "Reference Cell",
    "入力値": "Input",
}

STATUS_LABEL_EN = {

  "OK": "OK",
  "要確認": "Needs Review",
  "記入済み": "Completed",
  "未記入": "Not Entered",
  "対象外": "Not Applicable",
  "金額OK": "Amount OK",
  "添付済み": "Attached",
  "添付なし": "No Attachment",
  "画像は1枚のみ": "Only one image is allowed",
  "文字数不足": "Character Count Too Low",
  "文字数オーバー": "Character Limit Exceeded",
  "文字数超過": "Character Limit Exceeded",
  "文字数要確認": "Character Count Needs Review",
  "桁数間違い": "Invalid Number of Digits",
  "日付要確認": "Date Needs Review",
  "表記ゆれ（リストタブを再度ご確認ください）": "Inconsistent formatting (please check the List tab again)",
  "金額未達（10万未満）": "Amount Below Minimum",
  "金額超過（500万超）": "Amount Exceeds Maximum",
  "記入不足": "Incomplete",
  "メールアドレス形式要確認": "Email Format Needs Review",
  "未記入（カタカナ）": "Not Entered",
  "未記入（漢字）": "Not Entered"

}

DETAIL_ORDER = [
    ("-", "ファイル名"),
    ("研究計画調書_1枚目", "提出日"),
    ("研究計画調書_1枚目", "e-Rad 研究者番号"),
    ("研究計画調書_1枚目", "メールアドレス"),
    ("研究計画調書_1枚目", "研究代表者氏名"),
    ("研究計画調書_1枚目", "生年月日"),
    ("研究計画調書_1枚目", "e-Rad所属機関コード"),
    ("研究計画調書_1枚目", "所属機関"),
    ("研究計画調書_1枚目", "部局"),
    ("研究計画調書_1枚目", "職"),
    ("研究計画調書_1枚目", "所属機関の区分"),
    ("研究計画調書_1枚目", "応募者属性の区分"),
    ("研究計画調書_1枚目", "学生フラグ"),
    ("研究計画調書_1枚目", "研究領域"),
    ("研究計画調書_1枚目", "メインユースケース分類"),
    ("研究計画調書_1枚目", "メインユースケース分類（その他）"),
    ("研究計画調書_1枚目", "サブユースケース選択"),
    ("研究計画調書_1枚目", "研究課題名"),
    ("研究計画調書_1枚目", "AI活用度選択"),
    ("研究計画調書_1枚目", "現在の具体的な活用方法"),
    ("研究計画調書_1枚目", "研究経費"),
    ("研究計画調書_2枚目", "研究目的"),
    ("研究計画調書_2枚目", "研究方法"),
    ("研究計画調書_2枚目", "AI利活用の妥当性・実現可能性"),
    ("研究計画調書_2枚目", "達成目標"),
    ("研究計画調書_2枚目", "AI利活用のノウハウ抽出・共有の実現計画"),
    ("研究計画調書_2枚目", "成果の公開方針（任意）"),
    ("研究計画調書_2枚目", "研究業績等"),
    ("研究計画調書_2枚目", "画像の添付"),
    ("研究計画調書_3枚目", "設備備品費の明細"),
    ("研究計画調書_3枚目", "消耗品費の明細"),
    ("研究計画調書_3枚目", "設備備品費、消耗品費の必要性"),
    ("研究計画調書_3枚目", "謝金の明細"),
    ("研究計画調書_3枚目", "国内旅費の明細"),
    ("研究計画調書_3枚目", "外国旅費の明細"),
    ("研究計画調書_3枚目", "謝金、旅費の必要性"),
    ("研究計画調書_3枚目", "その他の明細"),
    ("研究計画調書_3枚目", "その他費用の必要性"),
    ("研究計画調書_4枚目", "費用詳細 API費用"),
    ("研究計画調書_4枚目", "費用詳細 計算資源費用 (クラウドGPU含む)"),
]
DETAIL_ORDER_MAP = {key: idx for idx, key in enumerate(DETAIL_ORDER)}


def get_major_item(sheet, item):
    """詳細チェックの大項目を返す。"""
    return MAJOR_ITEM_MAP.get(item, "-")


def detail_sort_key(detail):
    key = (detail.get("タブ名", ""), detail.get("項目", ""))
    return DETAIL_ORDER_MAP.get(key, len(DETAIL_ORDER))


def detail_row(file_path, sheet, item, status, cell_or_range="", value="", note=""):
    return {
        "ファイル名": safe_text_for_output(file_path.name),
        "タブ名": safe_text_for_output(sheet),
        "大項目": safe_text_for_output(get_major_item(sheet, item)),
        "項目": safe_text_for_output(item),
        "判定": safe_text_for_output(status),
        "参照セル": safe_text_for_output(cell_or_range),
        "入力値": safe_text_for_output(value),
        "補足": safe_text_for_output(note),
    }


def check_sheet_1(file_path, wb_values, wb_formula, details, language=LANG_JA):
    expected_sheet = get_expected_sheet_name(language, "sheet1")
    sheet_name = find_sheet_name_for_transfer(wb_values, expected_sheet)
    internal_sheet = get_internal_sheet_name("sheet1")
    if not sheet_name:
        return "要確認"
    get = lambda c: get_formula_aware_value(wb_values, wb_formula, sheet_name, c)

    def add(item, status, cell, value="", note=""):
        details.append(detail_row(file_path, internal_sheet, item, status, cell, value, note))

    v = get("C8")
    add("e-Rad 研究者番号", "記入済み" if has_input(v) else "未記入", "C8", v)

    v = get("C10")
    if not has_input(v):
        st = "未記入"
    elif re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]+", str(v).strip()):
        st = "記入済み"
    else:
        st = "メールアドレス形式要確認"
    add("メールアドレス", st, "C10", v)

    submission_year = get("H6")
    submission_month = get("J6")
    submission_day = get("L6")
    add(
        "提出日",
        judge_submission_date(submission_year, submission_month, submission_day),
        "H6/J6/L6",
        f"年：{safe_display_text(submission_year)} / 月：{safe_display_text(submission_month)} / 日：{safe_display_text(submission_day)}",
    )

    if language == LANG_EN:
        full_name = get("D12")
        add("研究代表者氏名", "記入済み" if has_input(full_name) else "未記入", "D12", full_name)
    else:
        kana = get("D12")
        kanji = get("D13")
        if has_input(kana) and has_input(kanji):
            st = "記入済み"
        elif not has_input(kana) and has_input(kanji):
            st = "未記入（カタカナ）"
        elif has_input(kana) and not has_input(kanji):
            st = "未記入（漢字）"
        else:
            st = "未記入"
        add("研究代表者氏名", st, "D12/D13", f"カタカナ={safe_display_text(kana)} / 漢字={safe_display_text(kanji)}")

    birth_year = get("C15")
    birth_month = get("F15")
    birth_day = get("H15")
    add(
        "生年月日",
        judge_birthdate(birth_year, birth_month, birth_day),
        "C15/F15/H15",
        f"年：{safe_display_text(birth_year)} / 月：{safe_display_text(birth_month)} / 日：{safe_display_text(birth_day)}",
    )

    v = get("C16")
    digits = to_digits_text(v)
    if not has_input(v):
        st = "未記入"
    elif len(digits) in {3, 4, 7, 10}:
        st = "記入済み"
    else:
        st = "桁数間違い"
    add("e-Rad所属機関コード", st, "C16", v)

    v = get("C18")
    add("所属機関", "記入済み" if has_input(v) else "未記入", "C18", v, "自由記述のため表記ゆれ判定なし")

    for item, cell in [("部局", "C19"), ("職", "C20")]:
        v = get(cell)
        add(item, "記入済み" if has_input(v) else "未記入", cell, v)

    choice_checks = [
        ("所属機関の区分", "C21", get_allowed_kubun(language)),
        ("応募者属性の区分", "C23", get_allowed_applicant_attributes(language)),
        ("研究領域", "C27", get_allowed_research_areas(language)),
        ("メインユースケース分類", "C29", get_allowed_main_use_cases(language)),
    ]
    main_use_case_value = ""
    applicant_attribute_value = ""
    for item, cell, allowed in choice_checks:
        v = get(cell)
        if item == "メインユースケース分類":
            main_use_case_value = v
        if item == "応募者属性の区分":
            applicant_attribute_value = v
        if not has_input(v):
            st = "未記入"
        elif is_allowed_choice(v, allowed):
            st = "記入済み"
        else:
            st = "表記ゆれ（リストタブを再度ご確認ください）"
        add(item, st, cell, v)

        if item == "応募者属性の区分":
            student_flag = get("L20")
            applicant_is_student = is_student_attribute(applicant_attribute_value, language)
            flag_is_y = is_y(student_flag)
            if applicant_is_student and not flag_is_y:
                flag_status = "応募者属性の区分（C23） 要確認"
            elif (not applicant_is_student) and flag_is_y:
                flag_status = "学生フラグ（L20） 要確認"
            else:
                flag_status = "記入済み"
            add("学生フラグ", flag_status, "L20", student_flag)

    v = get("C31")
    if is_main_use_case_other(main_use_case_value, language):
        st = "記入済み" if has_input(v) else "未記入"
    else:
        st = "対象外"
    add("メインユースケース分類（その他）", st, "C31", v)

    values = [get(c) for c in SUB_USE_CASE_CELLS]
    selected_labels = selected_right_labels(get, SUB_USE_CASE_CELLS, SUB_USE_CASE_LABEL_MAP)
    add("サブユースケース選択", y_selection_status(values), " / ".join(SUB_USE_CASE_CELLS), selected_labels, "空欄またはYのみ許容。全セル空欄は未記入")

    v = get("C35")
    add("研究課題名", "記入済み" if has_input(v) else "未記入", "C35", v)

    values = [get(c) for c in AI_USAGE_CELLS]
    selected_labels = selected_right_labels(get, AI_USAGE_CELLS, AI_USAGE_LABEL_MAP)
    add("AI活用度選択", y_selection_status(values), " / ".join(AI_USAGE_CELLS), selected_labels, "空欄またはYのみ許容。全セル空欄は未記入")

    v = get("C42")
    add("現在の具体的な活用方法", "記入済み" if has_input(v) else "未記入", "C42", v)

    v = get("D48")
    num = to_number(v)
    if num is None:
        st = "未記入"
    elif num < 100:
        st = "金額未達（10万未満）"
    elif num > 5000:
        st = "金額超過（500万超）"
    else:
        st = "金額OK"
    add("研究経費", st, "D48", v)

    return "OK" if all(is_ok_detail_status(d) for d in details if d["タブ名"] == internal_sheet) else "要確認"


def check_sheet_2(file_path, wb_values, wb_formula, details, language=LANG_JA):
    expected_sheet = get_expected_sheet_name(language, "sheet2")
    sheet_name = find_sheet_name_for_transfer(wb_values, expected_sheet)
    internal_sheet = get_internal_sheet_name("sheet2")
    if not sheet_name:
        return "要確認"
    get = lambda c: get_formula_aware_value(wb_values, wb_formula, sheet_name, c)

    text_count_rules = TEXT_COUNT_RULES_BY_LANGUAGE.get(language, TEXT_COUNT_RULES_BY_LANGUAGE[LANG_JA])
    count_label = "ワード数" if language == LANG_EN else "文字数"
    for item, rule in text_count_rules.items():
        row = rule["row"]
        count_cell = f"E{row}"
        count_value = get(count_cell)
        if has_input(count_value):
            # Excel の自動計算結果（数式キャッシュ）あり → 従来どおり E 列の値で判定
            status = judge_text_count_by_e_only(
                count_value,
                rule["min"],
                rule["max"],
                rule["optional"],
            )
            value = f"{count_label}：{safe_display_text(count_value)}"
        else:
            # 数式キャッシュが無い場合は、本文 D 列から同じ計算式で再計算する（結果は数式と一致）。
            body_value = get(f"D{row}")
            recomputed = recompute_count_for_sheet2(body_value, language)
            status = judge_text_count_by_e_only(
                recomputed,
                rule["min"],
                rule["max"],
                rule["optional"],
            )
            value = f"{count_label}：{recomputed}{RECALC_MARKER_JA}"
        details.append(detail_row(file_path, internal_sheet, item, status, count_cell, value))

    achievement_values = [get(f"D{r}") for r in range(14, 19)]
    achievement_count = sum(1 for v in achievement_values if has_input(v))
    details.append(detail_row(
        file_path,
        internal_sheet,
        "研究業績等",
        count_result(achievement_count),
        "D14:D18",
        f"{achievement_count}項目記入済み" if achievement_count > 0 else "",
        " / ".join(str(v) for v in achievement_values if has_input(v)),
    ))

    image_count = count_images_on_sheet_from_xml(file_path, sheet_name)
    if image_count == 0:
        image_status = "添付なし"
    elif image_count == 1:
        image_status = "添付済み"
    else:
        image_status = "画像は1枚のみ"
    details.append(detail_row(file_path, internal_sheet, "画像の添付", image_status, expected_sheet, "-"))

    return "OK" if all(
        is_ok_detail_status(d)
        for d in details if d["タブ名"] == internal_sheet
    ) else "要確認"


def check_sheet_3(file_path, wb_values, wb_formula, details, language=LANG_JA):
    expected_sheet = get_expected_sheet_name(language, "sheet3")
    sheet_name = find_sheet_name_for_transfer(wb_values, expected_sheet)
    internal_sheet = get_internal_sheet_name("sheet3")
    if not sheet_name:
        return "要確認"
    ws = wb_values[sheet_name]

    get = lambda c: get_formula_aware_value(wb_values, wb_formula, sheet_name, c)
    detail_statuses = {}
    detail_output_cells = {
        "設備備品費の明細": [f"D{r}" for r in range(11, 31)],
        "消耗品費の明細": [f"M{r}" for r in range(11, 31)],
        "謝金の明細": [f"D{r}" for r in range(40, 60)],
        "国内旅費の明細": [f"H{r}" for r in range(40, 60)],
        "外国旅費の明細": [f"M{r}" for r in range(40, 60)],
        "その他の明細": [f"D{r}" for r in range(69, 89)],
    }
    for item in ["設備備品費の明細", "消耗品費の明細", "謝金の明細", "国内旅費の明細", "外国旅費の明細", "その他の明細"]:
        rng = DETAIL_RANGES[item]
        status, value, complete_count, incomplete_rows = analyze_detail_range_rows(ws, rng)
        detail_statuses[item] = status
        output_value = display_values_from_cells(get, detail_output_cells[item])
        details.append(detail_row(file_path, internal_sheet, item, status, rng, output_value))

    necessity_logic = [
        ("設備備品費、消耗品費の必要性", "C34", ["設備備品費の明細", "消耗品費の明細"]),
        ("謝金、旅費の必要性", "C63", ["謝金の明細", "国内旅費の明細", "外国旅費の明細"]),
        ("その他費用の必要性", "C92", ["その他の明細"]),
    ]
    for item, cell, related_items in necessity_logic:
        value = get_formula_aware_value(wb_values, wb_formula, sheet_name, cell)
        if all(detail_statuses.get(x) == "未記入" for x in related_items):
            status = "対象外"
        else:
            status = "記入済み" if has_input(value) else "未記入"
        details.append(detail_row(file_path, internal_sheet, item, status, cell, value))

    return "OK" if all(
        is_ok_detail_status(d)
        for d in details if d["タブ名"] == internal_sheet
    ) else "要確認"


def check_sheet_4(file_path, wb_values, wb_formula, details, language=LANG_JA):
    expected_sheet = get_expected_sheet_name(language, "sheet4")
    sheet_name = find_sheet_name_for_transfer(wb_values, expected_sheet)
    internal_sheet = get_internal_sheet_name("sheet4")
    if not sheet_name:
        return "要確認"
    ws = wb_values[sheet_name]
    get = lambda c: get_formula_aware_value(wb_values, wb_formula, sheet_name, c)
    detail_output_cells = {
        "費用詳細 API費用": [f"D{r}" for r in range(9, 19)],
        "費用詳細 計算資源費用 (クラウドGPU含む)": [f"D{r}" for r in range(22, 32)],
    }
    for item in ["費用詳細 API費用", "費用詳細 計算資源費用 (クラウドGPU含む)"]:
        rng = DETAIL_RANGES[item]
        count = count_filled_rows(ws, rng)
        output_value = display_values_from_cells(get, detail_output_cells[item])
        details.append(detail_row(file_path, internal_sheet, item, count_result(count), rng, output_value))
    return "OK" if all(is_ok_detail_status(d) for d in details if d["タブ名"] == internal_sheet) else "要確認"


def process_one_file(file_path):
    input_base_dir = get_configured_input_base_dir()
    raw_file_path = Path(file_path)
    try:
        file_path = validate_excel_file_path(raw_file_path, input_base_dir, must_exist=True)
    except Exception as validation_error:
        details = []
        summary = {
            "ファイル名": safe_text_for_output(raw_file_path.name),
            "ファイルパス": safe_text_for_output(str(raw_file_path)),
            "言語": LANG_JA,
            "指定されたファイル名になっているか": "要確認",
            "タブ名を変更していないか": "要確認",
            "研究計画調書_1枚目": "要確認",
            "研究計画調書_2枚目": "要確認",
            "研究計画調書_3枚目": "要確認",
            "研究計画調書_4枚目": "要確認",
            "総合判定": "要確認",
            "要確認項目": "",
            "エラー内容": safe_text_for_output(f"ファイルパス検証エラー: {validation_error}"),
            "詳細チェックシート名": "",
        }
        details.append(detail_row(raw_file_path, "-", "ファイル名", "要確認", "", raw_file_path.name))
        return summary, details

    details = []
    language = LANG_JA
    summary = {
        "ファイル名": safe_text_for_output(file_path.name),
        "ファイルパス": safe_text_for_output(str(file_path)),
        "言語": language,
        "指定されたファイル名になっているか": "要確認",
        "タブ名を変更していないか": "OK",
        "研究計画調書_1枚目": "要確認",
        "研究計画調書_2枚目": "要確認",
        "研究計画調書_3枚目": "要確認",
        "研究計画調書_4枚目": "要確認",
        "総合判定": "要確認",
        "要確認項目": "",
        "エラー内容": "",
        "詳細チェックシート名": "",
    }

    wb_values = None
    wb_formula = None
    try:
        wb_values = load_workbook(str(file_path), data_only=True, read_only=False, keep_links=False, keep_vba=file_path.suffix.lower() == ".xlsm")
        wb_formula = load_workbook(str(file_path), data_only=False, read_only=False, keep_links=False, keep_vba=file_path.suffix.lower() == ".xlsm")
        language = detect_workbook_language(wb_values)
        summary["言語"] = language
        summary["指定されたファイル名になっているか"] = "OK" if judge_file_name(file_path, language) == "OK" else "要確認"

        sheet1_for_title = find_sheet_name_for_transfer(wb_values, get_expected_sheet_name(language, "sheet1"))
        if sheet1_for_title:
            d12_value = normalize_text(get_cell_value(wb_values[sheet1_for_title], "D12"))
            if d12_value:
                prefix = "Detail Check" if language == LANG_EN else "詳細チェック"
                summary["詳細チェックシート名"] = safe_text_for_output(f"{prefix}_{d12_value}")
    except Exception as e:
        for wb in (wb_values, wb_formula):
            try:
                if wb is not None:
                    wb.close()
            except Exception:
                pass
        summary["エラー内容"] = f"ファイルを開けません: {e}"
        summary["総合判定"] = "要確認"
        details.append(detail_row(
            file_path,
            "-",
            "ファイル名",
            summary["指定されたファイル名になっているか"],
            "",
            file_path.name,
        ))
        return summary, details

    details.append(detail_row(
        file_path,
        "-",
        "ファイル名",
        summary["指定されたファイル名になっているか"],
        "",
        file_path.name,
    ))

    try:
        wrong_tabs = []
        for sheet_key in SHEET_KEYS:
            expected = get_expected_sheet_name(language, sheet_key)
            actual = find_sheet_name_for_transfer(wb_values, expected)
            if not actual:
                wrong_tabs.append(f"{get_internal_sheet_name(sheet_key)}: シートなし")
            elif normalize_sheet_name_loose(actual) != normalize_sheet_name_loose(expected):
                wrong_tabs.append(f"{get_internal_sheet_name(sheet_key)}: シート名が異なります（実名: {safe_display_text(actual)}）")

        summary["タブ名を変更していないか"] = "OK" if not wrong_tabs else "要確認"
        if wrong_tabs:
            details.append(detail_row(file_path, "-", "タブ名", "要確認", "", " / ".join(wrong_tabs)))

        summary["研究計画調書_1枚目"] = check_sheet_1(file_path, wb_values, wb_formula, details, language)
        summary["研究計画調書_2枚目"] = check_sheet_2(file_path, wb_values, wb_formula, details, language)
        summary["研究計画調書_3枚目"] = check_sheet_3(file_path, wb_values, wb_formula, details, language)
        summary["研究計画調書_4枚目"] = check_sheet_4(file_path, wb_values, wb_formula, details, language)

        if any(RECALC_MARKER_JA in str(d.get("入力値", "")) for d in details):
            summary["エラー内容"] = (
                RECALC_SUMMARY_NOTE_JA
                if not summary["エラー内容"]
                else f"{RECALC_SUMMARY_NOTE_JA} / {summary['エラー内容']}"
            )

        ng = [
            d["項目"] for d in details
            if not is_ok_detail_status(d)
        ]
        summary["要確認項目"] = "；".join(dict.fromkeys(ng))
        summary["総合判定"] = "OK" if summary["指定されたファイル名になっているか"] == "OK" and summary["タブ名を変更していないか"] == "OK" and not ng else "要確認"
        return summary, details
    except Exception as e:
        summary["エラー内容"] = str(e)
        summary["総合判定"] = "要確認"
        return summary, details
    finally:
        wb_values.close()
        wb_formula.close()


def collect_excel_files(input_path):
    input_path = safe_resolve_path(input_path)
    input_base_dir = input_path.parent if input_path.suffix.lower() in ALLOWED_EXCEL_SUFFIXES else input_path
    if input_path.is_file():
        return [validate_excel_file_path(input_path, input_base_dir, must_exist=True)]
    if not input_path.is_dir():
        return []

    configured_output = validate_output_file_path(OUTPUT_FILE)
    output_prefixes = (
        "研究計画調書_セルフチェック結果",
        "Research_Plan_Self_Check_Result",
        "research_plan_self_check_result",
    )
    patterns = ["*.xlsx", "*.xlsm"]
    files = []
    for pattern in patterns:
        files.extend(input_path.rglob(pattern) if RECURSIVE else input_path.glob(pattern))

    safe_files = []
    for candidate in files:
        try:
            safe_candidate = validate_excel_file_path(candidate, input_base_dir, must_exist=True)
        except Exception:
            continue
        if safe_candidate.name.startswith("~$"):
            continue
        if any(safe_candidate.stem.startswith(prefix) for prefix in output_prefixes):
            continue
        if safe_candidate == configured_output:
            continue
        safe_files.append(safe_candidate)

    return sorted(safe_files, key=lambda x: str(x).lower())


# 出力Excel

def translate_sheet_label(value, output_language):
    if output_language == LANG_EN:
        return SHEET_LABEL_EN.get(value, value)
    return value


def translate_item_label(value, output_language):
    if output_language == LANG_EN:
        return ITEM_LABEL_EN.get(value, value)
    return value


def translate_category_label(value, output_language):
    if output_language == LANG_EN:
        return CATEGORY_LABEL_EN.get(value, value)
    return value


def translate_status(value, output_language):
    if output_language != LANG_EN:
        return value
    text = str(value)
    m = re.fullmatch(r"(\d+)項目記入済み", text)
    if m:
        return f"{m.group(1)} items completed"
    if text.startswith("応募者属性の区分（C23）"):
        return "Applicant attribute category (C23) needs review"
    if text.startswith("学生フラグ（L20）"):
        return "Student flag (L20) needs review"
    return STATUS_LABEL_EN.get(text, text)


def translate_free_text(value, output_language):
    if output_language != LANG_EN:
        return value
    if value is None:
        return ""
    text = str(value)
    for ja, en in SHEET_LABEL_EN.items():
        text = text.replace(ja, en)
    for ja, en in ITEM_LABEL_EN.items():
        text = text.replace(ja, en)
    replacements = {
        "年：": "Year: ",
        "月：": "Month: ",
        "日：": "Day: ",
        "文字数：": "Character count: ",
        "ワード数：": "Word count: ",
        "項目記入済み": " items completed",
        "記入不足行": "Incomplete rows",
        "シートなし": "Sheet missing",
        "シート名が異なります（実名:": "Sheet name differs (actual:",
        "ファイルを開けません:": "Could not open file:",
        "ファイルパス検証エラー:": "File path validation error:",
        "対象Excelが見つかりません:": "Target Excel file was not found:",
        "カタカナ=": "Katakana=",
        "漢字=": "Kanji=",
        "；": "; ",
        RECALC_MARKER_JA: RECALC_MARKER_EN,
        RECALC_SUMMARY_NOTE_JA: RECALC_SUMMARY_NOTE_EN,
    }
    for ja, en in replacements.items():
        text = text.replace(ja, en)
    text = re.sub(r"(\d+) items completed", r"\1 items completed", text)
    return text


def translate_review_items(value, output_language):
    if output_language != LANG_EN:
        return value
    parts = [p for p in str(value).split("；") if p]
    return "; ".join(translate_item_label(p, output_language) for p in parts)


def translate_summary_header(header, output_language):
    if output_language == LANG_EN:
        return SUMMARY_HEADER_EN.get(header, header)
    return header


def translate_detail_header(header, output_language):
    if output_language == LANG_EN:
        return DETAIL_HEADER_EN.get(header, header)
    return header


def translate_summary_value(value, header, output_language):
    if output_language != LANG_EN:
        return value
    if header in {"総合判定", "指定されたファイル名になっているか", "タブ名を変更していないか", "研究計画調書_1枚目", "研究計画調書_2枚目", "研究計画調書_3枚目", "研究計画調書_4枚目"}:
        return translate_status(value, output_language)
    if header == "要確認項目":
        return translate_review_items(value, output_language)
    if header == "エラー内容":
        return translate_free_text(value, output_language)
    return value


def translate_detail_value(value, header, output_language):
    if output_language != LANG_EN:
        return value
    if header == "タブ名":
        return translate_sheet_label(value, output_language)
    if header == "大項目":
        return translate_category_label(value, output_language)
    if header == "項目":
        return translate_item_label(value, output_language)
    if header == "判定":
        return translate_status(value, output_language)
    if header == "入力値":
        return translate_free_text(value, output_language)
    return value


def write_translated_sheet(ws, headers, rows, output_language, row_translator, header_translator):
    ws.append([safe_text_for_output(header_translator(h, output_language)) for h in headers])
    for row in rows:
        ws.append([safe_text_for_output(row_translator(row.get(h, ""), h, output_language)) for h in headers])


def get_header_col(header_cols, names):
    for name in names:
        if name in header_cols:
            return header_cols[name]
    return None


def style_workbook(wb):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    ok_fill = PatternFill("solid", fgColor="E2F0D9")
    ng_fill = PatternFill("solid", fgColor="FCE4D6")
    warn_fill = PatternFill("solid", fgColor="FFF2CC")
    gray_fill = PatternFill("solid", fgColor="D9D9D9")
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ok_values = {"OK", "記入済み", "添付済み", "対象外", "金額OK", "Completed", "Attached", "Not Applicable", "Amount OK"}
    ng_values = {
        "要確認", "未記入", "文字数不足", "文字数オーバー", "文字数超過", "文字数要確認", "桁数間違い", "NG",
        "金額オーバー", "金額アンダー", "金額未達（10万未満）", "金額超過（500万超）", "記入不足", "画像は1枚のみ", "日付要確認",
        "Needs Review", "Not Entered", "Count Too Low", "Count Exceeded", "Count Needs Review", "Invalid Number of Digits",
        "Amount Below Minimum", "Amount Exceeds Maximum", "Incomplete", "Only one image allowed", "Date Needs Review",
        "Email Format Needs Review", "Not Entered (Katakana)", "Not Entered (Kanji)",
    }
    gray_status_values = {"添付なし", "No Attachment"}
    optional_items = OK_UNWRITTEN_ITEMS | {ITEM_LABEL_EN.get(item, item) for item in OK_UNWRITTEN_ITEMS}

    for ws in wb.worksheets:
        is_detail_sheet = ws.title.startswith("詳細チェック") or ws.title.startswith("Detail Check")
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        header_cols = {ws.cell(1, col_idx).value: col_idx for col_idx in range(1, ws.max_column + 1)}
        item_col = get_header_col(header_cols, ["項目", "Item"])
        status_col = get_header_col(header_cols, ["判定", "Result"])
        input_col = get_header_col(header_cols, ["入力値", "Input Value"])

        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        for row in ws.iter_rows(min_row=2):
            item_value = ws.cell(row[0].row, item_col).value if is_detail_sheet and item_col else ""
            status_value = ws.cell(row[0].row, status_col).value if is_detail_sheet and status_col else ""
            detail_gray_status = (
                is_detail_sheet
                and (
                    (item_value in {"画像の添付", "Image Attachment"} and status_value in gray_status_values)
                    or (item_value in optional_items and status_value in {"未記入", "Not Entered"})
                )
            )
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border

                if is_detail_sheet and input_col and cell.column == input_col:
                    continue

                value = cell.value
                if detail_gray_status and status_col and cell.column == status_col:
                    cell.fill = gray_fill
                elif value in ok_values or (isinstance(value, str) and (value.endswith("項目記入済み") or value.endswith("items completed"))):
                    cell.fill = ok_fill
                elif value in ng_values:
                    cell.fill = ng_fill
                elif value in gray_status_values:
                    cell.fill = gray_fill if is_detail_sheet else ng_fill
                elif isinstance(value, str) and ("表記ゆれ" in value or "ご確認" in value or "inconsistency" in value or "needs review" in value.lower()):
                    cell.fill = warn_fill if "inconsistency" in value else ng_fill
                elif isinstance(value, str) and ("要確認" in value or "Needs Review" in value):
                    cell.fill = ng_fill

        for col_idx in range(1, ws.max_column + 1):
            header = ws.cell(1, col_idx).value or ""
            letter = get_column_letter(col_idx)
            if header in {"ファイルパス", "要確認項目", "入力値", "File Path", "Items Needing Review", "Input Value"}:
                width = 55
            elif header in {"ファイル名", "File Name"}:
                width = 34
            elif header in {"大項目", "Category"}:
                width = 28
            elif header in {"項目", "Item"}:
                width = 46
            elif header in {"判定", "総合判定", "Result", "Overall Result"}:
                width = 18
            elif header in {"参照セル", "Reference Cell"}:
                width = 22
            else:
                width = min(max(len(str(header)) + 4, 14), 40)
            ws.column_dimensions[letter].width = width
        ws.row_dimensions[1].height = 32


def safe_sheet_title(title, used_titles):
    invalid_chars = '[]:*?/\\'
    title_text = safe_display_text(title)
    title_text = re.sub(r"[\r\n\t]+", "_", title_text)
    safe = ''.join('_' if ch in invalid_chars else ch for ch in title_text)
    safe = safe.strip() or '詳細チェック'
    safe = safe[:31]
    base = safe
    idx = 2
    while safe in used_titles:
        suffix = f'_{idx}'
        safe = (base[:31 - len(suffix)] + suffix)[:31]
        idx += 1
    used_titles.add(safe)
    return safe


def determine_output_language(summaries):
    """英語版のみなら英語、混在または日本語を含む場合は日本語優先。"""
    if summaries and all(summary.get("言語") == LANG_EN for summary in summaries):
        return LANG_EN
    return LANG_JA


def build_output(summaries, details, output_file, output_language=LANG_JA):
    output_file = validate_output_file_path(output_file)
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary" if output_language == LANG_EN else "判定サマリー"
    summary_headers = [
        "ファイル名", "ファイルパス", "総合判定", "指定されたファイル名になっているか", "タブ名を変更していないか",
        "研究計画調書_1枚目", "研究計画調書_2枚目", "研究計画調書_3枚目", "研究計画調書_4枚目",
        "要確認項目", "エラー内容",
    ]
    write_translated_sheet(ws, summary_headers, summaries, output_language, translate_summary_value, translate_summary_header)

    detail_headers = ["ファイル名", "タブ名", "大項目", "項目", "判定", "参照セル", "入力値"]
    used_titles = {ws.title}
    for summary in summaries:
        file_name = summary.get("ファイル名", "")
        file_details = sorted([d for d in details if d.get("ファイル名") == file_name], key=detail_sort_key)
        if output_language == LANG_EN:
            preferred_title = summary.get("詳細チェックシート名") or f"Detail Check_{Path(str(file_name)).stem}"
            preferred_title = str(preferred_title).replace("詳細チェック", "Detail Check")
        else:
            preferred_title = summary.get("詳細チェックシート名") or f"詳細チェック_{Path(str(file_name)).stem}"
            preferred_title = str(preferred_title).replace("Detail Check", "詳細チェック")
        title = safe_sheet_title(preferred_title, used_titles)
        ws_detail = wb.create_sheet(title)
        write_translated_sheet(ws_detail, detail_headers, file_details, output_language, translate_detail_value, translate_detail_header)

    style_workbook(wb)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_file))


def run(input_path=None, output_file=None):
    """チェックを実行する。

    Notebook から呼ぶときは run(r"フォルダのパス") のように対象を渡せます。
    省略したときは、上部で設定した INPUT_PATH / OUTPUT_FILE を使います。
    出力した Microsoft Excel ファイルのパスを返します。
    """
    global INPUT_PATH, OUTPUT_FILE
    saved_input, saved_output = INPUT_PATH, OUTPUT_FILE
    try:
        if input_path is not None:
            INPUT_PATH = Path(input_path)
        if output_file is not None:
            OUTPUT_FILE = Path(output_file)
        files = collect_excel_files(INPUT_PATH)
        if not files:
            raise FileNotFoundError(f"対象Excelが見つかりません: {INPUT_PATH}")
        summaries = []
        all_details = []
        print(f"対象Excel: {len(files)}件")
        for idx, file_path in enumerate(files, start=1):
            print(f"処理中 {idx}/{len(files)}: {safe_display_text(file_path.name)}")
            summary, details = process_one_file(file_path)
            summaries.append(summary)
            all_details.extend(details)
        output_language = determine_output_language(summaries)
        result_path = get_timestamped_output_file(output_language)
        build_output(summaries, all_details, result_path, output_language)
        print(f"完了: {safe_display_text(result_path)}")
        return result_path
    finally:
        INPUT_PATH, OUTPUT_FILE = saved_input, saved_output


if __name__ == "__main__":
    run()
